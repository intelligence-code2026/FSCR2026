'''本程序用于计算补充测试集合的指标，使用的优化后模型与main_model中的模型相同，主要是验证模型的泛化能力。
修改epoch_num_cv值可以加载交叉验证中得到的不同轮次的优化模型。'''
''''This program is designed to calculate metrics for the supplementary test set.
It utilizes the optimized model, which is identical to the one in main_model, primarily to verify the model's generalization capability.
By modifying the epoch_num_cv value, you can load different rounds of the optimized models obtained during cross-validation.'''

# In[0]进行基本设置，定义基本函数
# Configure basic settings and define fundamental functions.
import random
import time
import openpyxl
import dspy
import os
import re
import numpy as np
from openai import OpenAI
from sklearn.metrics.pairwise import cosine_similarity
import copy, pickle
import self_func
import sys

num_cores = os.cpu_count()
print(f"CPU核心数: {num_cores}")

# =============================================================================
# epoch_num_cv是交叉验证轮数，可以修改！！！！！！！！！！！！！！！！！！！！！！！！！！
# epoch_num_cv: Number of cross-validation folds (modifiable)
# =============================================================================
epoch_num_cv=1      #这个值最大是5，每轮验证完要修改这个值
assert 1 <= epoch_num_cv <= 5, "epoch_num_cv 应为 1 到 5 的整数"

# =============================================================================
# num_threads_lim是多线程计算的线程上限，可以修改！！！！！！！！！！！！！！！！！！
# num_threads_lim: Maximum thread count for parallel computation (modifiable)
# =============================================================================
num_threads_lim = 8

# =============================================================================
# num_query是生成的查询案例数量，包括训练集和验证集，可以修改！！！！！！！！！！！！！！！
# num_query: Total count of generated queries for train/val sets (modifiable)
# =============================================================================
num_query = 800

# =============================================================================
# num_testset是生成的测试集中数据量，可以修改！！！！！！！！！！！！！！！！！！！！！！！
# num_testset: Size of the generated test dataset (modifiable)
# =============================================================================
num_testset=200

# =============================================================================
# num_candidate_case是是生成查询案例时，引用的候选案例数量，可以修改！！！！！！！！
# num_candidate_case: Count of candidate cases used for query generation (modifiable)
# =============================================================================
num_candidate_case=3

# =============================================================================
# num_summary是生成查询案例时，每个候选案例引用的summary数量，可以修改！！！！！！！
# num_summary: Count of summaries per candidate case for query generation (modifiable)
# =============================================================================
num_summary = 2

# =============================================================================
# num_candidate是候选案例数量，可以修改！！！！！！！！！！！！！！！！！！！！！！！
# num_candidate: Total count of candidate cases (modifiable)
# =============================================================================
num_candidate = 37

# =============================================================================
# max_tokens_lm是大模型输出的tokens限值，可以修改！！！！！！！！！！！！！！！
# max_tokens_lm: Max output tokens for the LLM (modifiable)
# =============================================================================
max_tokens_lm = None

# =============================================================================
# similar_threshold_value是细粒度类案检索相似的阈值，可以修改！！！！！！！！！！！
# similar_threshold_value: Threshold for fine-grained case similarity search (modifiable)
# =============================================================================
# similar_threshold_value = 0.74    	      #目前similar_threshold_value=0.89为最优
DEFAULT_SIMILAR_THRESHOLD = 0.9

# =============================================================================
# num_candidate_selected_value是在build_dict2_from_query_chains函数中，
# 根据与query的相似度筛选出candidate的数量，可以修改！！！！！！！！！！！！！！！！
# num_candidate_selected_value: Count of candidates filtered by query similarity 
# in build_dict2_from_query_chains(modifiable)
# =============================================================================
num_candidate_selected_value=37

# =============================================================================
# auto_option是优化参数设置，可以选择None、"light"、"medium"、"heavy"！！！！！！！
# auto_option: Optimization level (None, 'light', 'medium', 'heavy')
# =============================================================================
auto_option = "heavy"

# 并发线程数（实际用于 ThreadPoolExecutor）
# Concurrent thread count (used by ThreadPoolExecutor)
num_threads = min(num_threads_lim, num_cores)

# 读取 API Key
# Load API Key
try:
    xlsx_api_key = openpyxl.load_workbook(r'..\..\api_key.xlsx')
    sheet_api_key = xlsx_api_key['Sheet1']
    api_key_deepseek = sheet_api_key['B7'].value
    api_key = sheet_api_key['B' + str(2)].value
except Exception as e:
    print('未找到llm的api_key！！！')
    print(f"错误详情: {e}")
    sys.exit(1)

try:             
    dict_case_text_pickle=r'.\data\dspy_opted_model\dict_case_text.pkl'
    with open(dict_case_text_pickle,'rb') as f_obj9:
         dict_case_text=pickle.load(f_obj9)     
    
except FileNotFoundError:    
    print('加载dict_case_text时出错未找到该文件，要重新生成。')
    # 读取候选案例文本
    xlsx_candidation = openpyxl.load_workbook(r'.\data\candidate_case2.xlsx')
    sheet_candidate = xlsx_candidation['Sheet1']
    total_rows = sheet_candidate.max_row
    dict_case_text = {}
    for k in list(range(2, total_rows + 1))[:num_candidate]:
        dict_case_text[k] = sheet_candidate['J' + str(k)].value
    xlsx_candidation.close()                 
    dict_case_text_pickle = r'.\data\dspy_opted_model\dict_case_text.pkl'
    with open(dict_case_text_pickle,'wb') as f_obj8:
          pickle.dump(dict_case_text, f_obj8)      

print(f'由xlsx生成的dict_case_text包含的数据共{len(dict_case_text)}条。')



# 配置 LLM
# Configure LLM
deepseek_lm = dspy.LM("openai/qwen-turbo", api_key=api_key, max_tokens=max_tokens_lm,api_base="https://dashscope.aliyuncs.com/compatible-mode/v1", temperature=0.1)

lm_high_temperature = dspy.LM("openai/deepseek-chat", api_key=api_key_deepseek, max_tokens=max_tokens_lm,api_base="https://api.deepseek.com",temperature=0.3)

lm = dspy.LM("openai/deepseek-chat", api_key=api_key_deepseek, max_tokens=max_tokens_lm,api_base="https://api.deepseek.com",temperature=0.2)

dspy.configure(lm=lm)

# clien是词嵌入大模型
# client: The large language model used for text embeddings
client = OpenAI(
   api_key=sheet_api_key['B'+str(2)].value,  
   base_url="https://dashscope.aliyuncs.com/compatible-mode/v1"  )

# ========================================
# 初始提示词（会被优化）
# Initial prompt (optimizable
# ========================================

file_path_summary = r'.\data\prompt\detail_summary_prompt_opt0-2.txt'    #读取原始的提示词文本文件
with open(file_path_summary, 'r', encoding='utf-8') as file:       
      detail_summary_prompt = file.read()      
PROMPT_SUMMARY = detail_summary_prompt      


# =============================================================================
# PROMPT_SUMMARY_CANDIDATE是用于抽取候选案例摘要的提示词
# PROMPT_SUMMARY_CANDIDATE: Prompt for extracting candidate case key circumstances
# =============================================================================
file_path_summary_candidate = r'.\data\prompt\detail_summary_prompt_opt0-3.txt'    #读取原始的提示词文本文件
with open(file_path_summary_candidate, 'r', encoding='utf-8') as file_summary_candidate:       
      detail_summary_prompt_candidate = file_summary_candidate.read()      
PROMPT_SUMMARY_CANDIDATE = detail_summary_prompt_candidate     

PROMPT_ACTION_CHAIN = "提示词2：请从下列摘要中提取关键行为链（用'->'连接每个动作），每行对应一条摘要的行为链，若一条摘要有多条行为链用;分开。输出必须为中文。"

file_path_compose = r'.\data\prompt\compose_prompt.txt'    #读取提示词文本文件
with open(file_path_compose, 'r', encoding='utf-8') as file_compose:       
      compose_prompt = file_compose.read()
PROMPT_COMPOSE_QUERY = compose_prompt     

PROMPT_JUDGE_QUERY = '以下文本中，命名为chains的文字是若干要点（用;分隔各要点），命名为query的文字是一个法律案例，你检查一下query是否具备以下要求：1.案情完整\n 2.逻辑无错误\n 3.包括了chain的全部情节，当事人名称、地点名称、时间可与chain不同。\n如果具备上述的全部要求，则输出标志语：yes，如果不完全具备上述的全部要求，输出标志语：no，标志语后面输出判断理由，标志语和判断理由之间用,分割。'   


def make_signature(input_name: str, output_name: str, prompt: str):
    Sig = type(
        "Sig",
        (dspy.Signature,),
        {
            "__annotations__": {input_name: str, output_name: str},
            input_name: dspy.InputField(),
            output_name: dspy.OutputField(),
        }
    )
    Sig.__doc__ = prompt
    return Sig


summarizer = dspy.Predict(make_signature("text", "summary", PROMPT_SUMMARY))

summarizer_candidate = dspy.Predict(make_signature("text", "summary", PROMPT_SUMMARY_CANDIDATE))

action_chain_extractor = dspy.Predict(make_signature("summary", "action_chain", PROMPT_ACTION_CHAIN))
composer = dspy.Predict(make_signature("action_chains", "composed_text", PROMPT_COMPOSE_QUERY))
judge = dspy.Predict(make_signature("composed_text", "judge_result", PROMPT_JUDGE_QUERY))

# ========================================
# 工具函数 (MODIFIED: 接收 Predictor 参数)
# Utility function (MODIFIED: now takes a Predictor argument)
# ========================================
def safe_split(s:str):
    # 步骤1：合并多个连续的 # 为一个
    s_clean = re.sub(r'#{2,}', '#', s)
    
    # 步骤2：按 # 分割，并去除空白项（可选）
    result = [x.strip() for x in s_clean.split('#') if x.strip()]
    
    return result

def call_summarizer(summarizer, text: str)->dict:
    dict_summary_temp={}
    p = summarizer(text=text) # **使用传入的参数**
    summary_text = getattr(p, "summary", "")
    try:
        list_summary=safe_split(summary_text)
    except:
        list_summary=[summary_text]
    for i in range(len(list_summary)):
        dict_summary_temp[i]=list_summary[i]
    return dict_summary_temp

def call_summarizer_candidate(summarizer_candidate, text: str)->dict:
    dict_summary_temp={}
    p = summarizer_candidate(text=text) # **使用传入的参数**
    summary_text = getattr(p, "summary", "")
    try:
        list_summary=safe_split(summary_text)
    except:
        list_summary=[summary_text]
    for i in range(len(list_summary)):
        dict_summary_temp[i]=list_summary[i]
    return dict_summary_temp

def call_action_extractor(action_chain_extractor, dict_summary_temp: dict)->dict:
    dict_chain_temp={}
    for j in dict_summary_temp.keys():
        try:
            p = action_chain_extractor(summary=dict_summary_temp[j]) # **使用传入的参数**
            chain_gen = getattr(p, "action_chain", "")
            # if isinstance(chain_gen, str) and chain_gen.strip():
            if chain_gen and chain_gen.strip():  
                dict_chain_temp[j]=chain_gen
            else:
                continue
        except Exception as e:
            continue       
    return dict_chain_temp

def call_composer(composer, dict_chain:dict)->str:
    num_factor=0
    input_text=''
    for key1 in dict_chain.keys():
        for key2 in dict_chain[key1].keys():
            num_factor+=1
            factor_text=dict_chain[key1][key2]
            input_text+=f'{num_factor}.{factor_text};'
    with dspy.context(lm=lm_high_temperature):
        p = composer(action_chains=input_text) 
    return getattr(p, "composed_text", ""), input_text

def call_judge(judge, query: str, chains:str, use_deepseek=True)->str:
    # 使用 deepseek_lm 检查 query 是否包含所有行为链且逻辑一致
    input_text=f'chains是{chains}。query是{query}'
    if use_deepseek:
        with dspy.context(lm=deepseek_lm):
            p = judge(composed_text=input_text) # **使用传入的参数**
    else:
        p = judge(composed_text=input_text) # **使用传入的参数**
    return getattr(p, "judge_result", "")

def embedding_vector(sentence:str,client=client)->list:     
    completion = client.embeddings.create(
        model="text-embedding-v3",
        input=sentence,
        dimensions=1024,
        encoding_format="float"
    )
    return  completion.data[0].embedding

def similiar_score(embedding_vec_query, embedding_vec_candidate)->float:
    # 确保输入是numpy数组（如果你的输入已经是数组，可以省略这步）
    vec1 = np.array(embedding_vec_query).reshape(1, -1)  # 强制转为2D: (1, n)
    vec2 = np.array(embedding_vec_candidate).reshape(1, -1)  # 强制转为2D: (1, n)
    
    # 计算余弦相似度（结果是一个1x1矩阵）
    similarity_matrix = cosine_similarity(vec1, vec2)
    
    # 返回标量值（提取矩阵中的单个数值）
    return similarity_matrix[0][0]

def generate_query_and_validate(composer, judge, dict_chain, dict1, max_retries=5)->dict:
    """
    用 lm 生成 query，并用 deepseek_lm 判定是否包含全部链且逻辑一致；
    若不符合则重试。最多尝试 max_retries + 1 次（含第一次）。
    """
    for attempt in range(max_retries + 1):
        try:
            query_text,chains = call_composer(composer,dict_chain)
            # print(f'***生成的query是:{query_text}')
            # print(f'***生成的chains是:{chains}')
            
            result_check = call_judge(judge, query_text, chains, use_deepseek=True)
            
            # result_check='yes'
            
            
            print(f'***对query检查结果是:{result_check}')
            if 'yes' in result_check.strip().lower():
                return {'text': query_text, 'label': dict1}
        except Exception as e:
            print(f"第 {attempt + 1} 次调用失败: {e}")
        
        # 如果不是最后一次尝试，等待后继续
        if attempt < max_retries:
            time.sleep(2)    
    # 所有尝试都失败
    print(f'共进行了 {max_retries + 1} 次查询生成，均失败。')
    return {'text': 'failed', 'label': dict1}



def process_query_extract(summarizer, dict_query)->dict:    
    """对 query 用提示词1和提示词2再次抽取摘要与行为链"""
    try:
        dict_summary_temp = call_summarizer(summarizer, dict_query['text']) 
        return {'chain':dict_summary_temp,'label':dict_query['label']}
    except Exception as e:        
        print(f"生成query的dict_summary时，调用失败: {e}")
        return 'gen_query_summary_failed'   
    
    

def process_candidate_extract(summarizer_candidate, action_chain_extractor, dict_case_text):
    """对 candidate 用提示词1和提示词2抽取摘要与行为链，并获得词向量"""
    dict_candidate_summary={}
    dict_candidate_chain={}
    dict_candidate_embedding={}
    for key_candidate in dict_case_text.keys():
        try:
            dict_summary_candidate_temp = call_summarizer_candidate(summarizer_candidate, dict_case_text[key_candidate])
            dict_candidate_summary[key_candidate]=dict_summary_candidate_temp
        except Exception as e:        
            print(f"生成第{key_candidate}个候选案例的摘要时，调用失败: {e}")
            continue
    for key_candidate_summary in dict_candidate_summary.keys():
        sub_dict_candidate_chain={}
        sub_dict_candidate_embedding={}
        for key_candidate_summary_sub in dict_candidate_summary[key_candidate_summary].keys():
            try:
                chain_extracted_candidate=action_chain_extractor(
                    summary=dict_candidate_summary[key_candidate_summary][key_candidate_summary_sub])   #获取某条summary的行为链
                chain_gen_candidate = getattr(chain_extracted_candidate, "action_chain", "")
                # if isinstance(chain_gen_candidate,str) and chain_gen_candidate.strip(): 
                if chain_gen_candidate and chain_gen_candidate.strip():          
                    try:                                               
                        sub_dict_candidate_embedding[key_candidate_summary_sub]=embedding_vector(dict_candidate_summary[key_candidate_summary][key_candidate_summary_sub])
                        sub_dict_candidate_chain[key_candidate_summary_sub]=chain_gen_candidate
                    except Exception as e:
                        print(f'第{key_candidate_summary}个候选案例的第{key_candidate_summary_sub}摘要的行为链词向量生成失败：{e}')
                        continue
                else:
                    continue
            except Exception as e: 
                print(f"生成第{key_candidate_summary}个候选案例的第{key_candidate_summary_sub}个行为链时，调用失败: {e}")
                continue
        dict_candidate_chain[key_candidate_summary]=sub_dict_candidate_chain
        dict_candidate_embedding[key_candidate_summary]=sub_dict_candidate_embedding
    
    return dict_candidate_summary,dict_candidate_chain,dict_candidate_embedding
       

def build_dict2_from_query_chains(embedding_dict_query, 
                                  embedding_dict_candidate_temp,
                                  num_candidate_selected=num_candidate_selected_value ,
                                  similar_threshold=DEFAULT_SIMILAR_THRESHOLD)->dict:
    
    list_key_candidate=self_func.rank_by_bidirectional_similarity(embedding_dict_query, embedding_dict_candidate_temp,num_candidate_selected)    
    embedding_dict_candidate={}
    for k in list_key_candidate:
        embedding_dict_candidate[k]=embedding_dict_candidate_temp[k]
    print(f'embedding_dict_candidate中有{len(embedding_dict_candidate)}个字典。')
    
    for key_query in embedding_dict_query.keys():    
        similar_dict_key_query={}    
        for q in embedding_dict_query[key_query].keys():
            similar_dict_key_candidate={}           
            for k in embedding_dict_candidate.keys():      # 仅加载用于生成查询案例的候选案例 
                similar_dict_candidate_k={}         
                for c in embedding_dict_candidate[k].keys():
                    if isinstance(embedding_dict_candidate[k][c], list) and isinstance(embedding_dict_query[key_query][q], list):
                        similiar_value=similiar_score(embedding_dict_query[key_query][q],embedding_dict_candidate[k][c])
                    else :
                        similiar_value=0
                    similar_dict_candidate_k[c]=copy.deepcopy(similiar_value)                    
                similar_dict_key_candidate[k]= copy.deepcopy(similar_dict_candidate_k)      
            similar_dict_key_query[q]= copy.deepcopy(similar_dict_key_candidate)  
                 
        # =============================================================================
        # 以下程序可以根据similar_threshold值求得与查询案例各条摘要具有最大相似度的查询案例摘要的编号
        # Finds the indices of query case summaries with the highest similarity to each candidate summary, subject to similar_threshold
        # =============================================================================
        
        max_similar_dict_key_query={}                
        for q2 in similar_dict_key_query.keys():
            max_condidate_of_condidate_case={}
            
            for k2 in similar_dict_key_query[q2].keys():
                condidate_of_condidate_case={}
               
                for c2 in similar_dict_key_query[q2][k2].keys():           
                    if similar_dict_key_query[q2][k2][c2]>similar_threshold:
                        condidate_of_condidate_case[c2]=similar_dict_key_query[q2][k2][c2]
                if condidate_of_condidate_case!={}:       
                    max_key_condidate = max(condidate_of_condidate_case, key= condidate_of_condidate_case.get)   #取得具有最大相似度的键  
                    
                    # max_condidate_of_condidate_case是一个字典，键为第k2个候选案例，值为与查询案例第q2条案情摘要具有最大相似度的候选案例的摘要编号    
                    max_condidate_of_condidate_case[k2]=max_key_condidate
                                                
                else:
                #     max_key_condidate=None
                    continue
                
            if max_condidate_of_condidate_case!={}:  
                
                if len(max_condidate_of_condidate_case)>1:
                    max_similar_dict={}
                    similiar_dict={}
                    for key in max_condidate_of_condidate_case.keys():
                        similiar_dict[key]=similar_dict_key_query[q2][key][max_condidate_of_condidate_case[key]]               
                    max_key = max(similiar_dict, key=similiar_dict.get)  # 找出最大值对应的键
                    max_similar_dict[max_key]=max_condidate_of_condidate_case[max_key]
                elif len(max_condidate_of_condidate_case)==1:
                    max_similar_dict=max_condidate_of_condidate_case
                
                max_similar_dict_key_query[q2]=max_similar_dict  
            else:
                continue                        
        query_pre_dict = {}
        
        # 遍历原始字典
        # Iterate over the original dictionary
        for outer_key, inner_dict in max_similar_dict_key_query.items():
            for inner_key, value in inner_dict.items():
                if inner_key not in query_pre_dict:
                    query_pre_dict[inner_key] = set()  # 使用集合自动去重
                query_pre_dict[inner_key].add(value)  # 添加值到集合               
        dict2 = {k: list(v) for k, v in query_pre_dict.items()}
        print(dict2)
    return dict2


def build_dict2_from_query_chains_2(embedding_dict_query, 
                                    embedding_dict_candidate_temp,
                                    num_candidate_selected=num_candidate_selected_value ,
                                    similar_threshold=DEFAULT_SIMILAR_THRESHOLD)-> tuple[dict, dict]:
    
    '''输出的output键是embedding_dict_query[0]对应的键，值是一个元组，
    该元组第一个元素是 embedding_dict_candidate_temp的键（即第几个候选案例）
    该元组第二个元素是embedding_dict_candidate_temp中作为值的字典的键（即某个候选案例的第几个摘要或者链）.
    输出的dict2是一个字典，是对dict1的预测值。'''
    
    list_key_candidate=self_func.rank_by_bidirectional_similarity(embedding_dict_query, embedding_dict_candidate_temp,num_candidate_selected)    
    embedding_dict_candidate={}
    for k in list_key_candidate:
        embedding_dict_candidate[k]=embedding_dict_candidate_temp[k]
    print(f'embedding_dict_candidate中有{len(embedding_dict_candidate)}个字典。')
        
    num_candidate_embedding=0
    list_candidate_embedding=[]
    dict_site={}
    for k1 in embedding_dict_candidate.keys():        
        for k2 in embedding_dict_candidate[k1].keys():
            list_candidate_embedding.append(embedding_dict_candidate[k1][k2])
            dict_site[num_candidate_embedding]=(k1,k2)
            num_candidate_embedding+=1
    
    k5=0
    site_query={}
    for k4 in embedding_dict_query[0].keys():
        site_query[k5]=k4
        k5+=1
        
    array_candidate_embedding=np.array(list_candidate_embedding)
    embedding_dict_query_temp = list(embedding_dict_query.values())[0]
    array_query_embedding = np.array(list(embedding_dict_query_temp.values()))
    qc_sim_embedding = cosine_similarity(array_query_embedding, array_candidate_embedding)
    print(f'embedding_dict_query共有{len(embedding_dict_query[0])}个词向量，qc_sim_embedding的形状是{qc_sim_embedding.shape}')
    mask = qc_sim_embedding > similar_threshold
        
    # 如果某行没有满足条件的元素，则该行最大值设为 -inf，避免干扰 argmax
    qc_masked = np.where(mask, qc_sim_embedding, -np.inf)

    # 每行第一个最大值的列索引（argmax 默认返回第一个最大值的位置）
    col_indices = np.argmax(qc_masked, axis=1)

    # 判断每行是否至少有一个元素 > threshold
    valid_rows = np.any(mask, axis=1)

    # 获取有效的行索引
    row_indices = np.where(valid_rows)[0]
    col_indices = col_indices[valid_rows]

    # 转换为 (行, 列) 元组的列表,result是一个列表，列表每个元素都是一个元组，
    # 元组第一个值是选中的qc_sim_embedding中元素的行号，元组第二个值是列号
    result = [(int(row), int(col)) for row, col in zip(row_indices, col_indices)]
    output={}
    for k3 in range(len(result)):
        output[site_query[result[k3][0]]]= dict_site[result[k3][1]]
        
    dict2 = {}
    for x, y in output.values():
        dict2.setdefault(x, []).append(y)
   
    return output,dict2


# ========================================
# pipeline 函数 (MODIFIED: 传递 Predictor 实例)
# Pipeline function (MODIFIED: passes a Predictor instance)
# ========================================
def pipeline_func(dict_query, 
                  embedding_dict_candidate, 
                  summarizer,
                  similar_threshold=DEFAULT_SIMILAR_THRESHOLD
                  ):
                   
    #  对 query 抽取摘要和行为链
    embedding_dict_query_temp={}
    if dict_query['text']!='failed':
        query_chains = process_query_extract(summarizer, dict_query) # **传递 Predictor**
        if isinstance(query_chains, dict):
            for key_query_chains in query_chains['chain'].keys():
                if query_chains['chain'][key_query_chains]:
                    embedding_dict_query_temp[key_query_chains]=embedding_vector(query_chains['chain'][key_query_chains])
                else:
                    continue
        else:
            print(f'query_chain不是个dict，而是{query_chains},出错了！！！！！')
            return {"dict1": {},"dict2": {}}
    else:
        print(f'当前的dict_query是{dict_query},出错了！！！！！')
        return {"dict1": {},"dict2": {}}
    
    embedding_dict_query={0:embedding_dict_query_temp}                              
        
    try:
        if embedding_dict_query[0]:
            query2candidate,dict2 = build_dict2_from_query_chains_2(
                                    embedding_dict_query, 
                                    embedding_dict_candidate,
                                    similar_threshold=similar_threshold  # <-- 传入阈值
                                                                    )
        else:
            print(f'当前的embedding_dict_query[0]是{embedding_dict_query[0]},出错了！！！！！')
            dict2 = {}
            query2candidate={}
    except Exception as e:
        print(f"构建dict2失败: {e}")
        dict2 = {}
        query2candidate={}
    return {"dict1": dict_query['label'],"dict2": dict2, "query2candidate":query2candidate,"query_chains":query_chains }

def my_metric(example, pred, trace=None):
    
    if pred.dict2=={} or pred.dict1=={}:
        
        return -1
    else:
        precision,recall,f1=self_func.metrics_func(pred.dict2, pred.dict1)     
        return float(f1)
    
def evalution_metric(example, pred, trace=None):
    if pred.dict2=={} or pred.dict1=={}:
        return -1
    else:
        precision,recall,f1=self_func.metrics_func(pred.dict2, pred.dict1)     
        return [precision,recall,f1]    
print('当前运行的单元格是In[0]，已经完成。')    
# In[1]生成训练集，验证集
# Generate training and validation sets

# In[2]进行提示词优化
# Optimize prompts

# In[3] optuna优化超参数
# Optimize hyperparameters with Optuna

# In[4] 计算验证集得分
# # Compute validation set metrics

# In[5.1]测试集的生成 
# Test set generation 

dict_candidate_summary_pickle=r'.\data\dspy_opted_model\dict_candidate_summary.pkl'
with open(dict_candidate_summary_pickle,'rb') as f_obj1:
     dict_candidate_summary=pickle.load(f_obj1) 
     
dict_candidate_chain_pickle=r'.\data\dspy_opted_model\dict_candidate_chain.pkl'
with open(dict_candidate_chain_pickle,'rb') as f_obj2:
     dict_candidate_chain=pickle.load(f_obj2) 
     
embedding_dict_candidate_pickle=r'.\data\dspy_opted_model\embedding_dict_candidate.pkl'
with open(embedding_dict_candidate_pickle,'rb') as f_obj3:
     embedding_dict_candidate=pickle.load(f_obj3)              


try:
    list_query_test_pickle=r'.\data\dspy_opted_model\list_query_test.pkl'
    with open(list_query_test_pickle,'rb') as f_obj12:
         list_query_test=pickle.load(f_obj12)    
    assert len(list_query_test) == num_testset,f"加载的测试集案例数量为 {len(list_query_test)} 与预期 {num_testset} 不符"
except (FileNotFoundError, EOFError, pickle.UnpicklingError, AssertionError) as e_list_query_test:    
    print('加载测试集时出错',e_list_query_test)

    if isinstance(e_list_query_test, FileNotFoundError):
        print('测试集文件未找到，需要重新生成。')
        list_query_test=[]
    elif isinstance(e_list_query_test, AssertionError):
        print(f'测试集加载的查询案例不够，需要补充生成{num_testset-len(list_query_test)}条案例数据。')
        
    add_num_query_test=num_testset-len(list_query_test)
    for _ in range(add_num_query_test): 
         dict1={}
         dict_chain={}
         list_candidate=random.sample(list(embedding_dict_candidate.keys()), num_candidate_case)  
         for k_dict in list_candidate:

# =============================================================================
# 从每个候选案例的行为链词嵌入字典中随机抽取键构建dict1
# Construct dict1 by randomly sampling keys from the behavior chain word embedding dictionary of each candidate case
# =============================================================================
             dict_summaries = embedding_dict_candidate[k_dict]
             list_summary_label=random.sample(list(dict_summaries.keys()), min(num_summary,len(dict_summaries.keys())))  
        
             dict1[k_dict]=copy.deepcopy(list_summary_label)
             sub_dict_chain={}
             for key_list in list_summary_label:
                       
                 try:           
                     
                     sub_dict_chain[key_list]=dict_candidate_summary[k_dict][key_list]
                 except Exception as e:
                     print(f'在第{k_dict}个候选案例的第{key_list}摘要调用对应行为链出错',e)
                     print(f'第{k_dict}个候选案例的摘要键列表为{list_summary_label}，行为链键列表为{list(dict_candidate_chain[k_dict])}')
             dict_chain[k_dict]=sub_dict_chain            
         dict_query=generate_query_and_validate(composer, judge, dict_chain, dict1, max_retries=5)
         if dict_query['text']=='failed':
              continue
         list_query_test.append(dict_query) 

    list_query_test_pickle=r'.\data\dspy_opted_model\list_query_test.pkl'  
    with open(list_query_test_pickle,'wb') as f_obj13:
         pickle.dump(list_query_test, f_obj13)           
                  
testset = []                             #dataset是全部数据的集合
for query_test in list_query_test:   
    example = dspy.Example(dict_query=query_test, label=query_test['label']).with_inputs("dict_query")
    testset.append(example)
print('当前运行的单元格是In[5.1]，已经完成。')    

# In[5.2] 加载最初优化前的提示词
# Load the original prompt before optimization
file_path_summary = r'.\data\prompt\detail_summary_prompt_opt0-2.txt'    
with open(file_path_summary, 'r', encoding='utf-8') as file1:       
      detail_summary_prompt_1 = file1.read()      
PROMPT_SUMMARY_1 = detail_summary_prompt_1 
summarizer_1 = dspy.Predict(make_signature("text", "summary", PROMPT_SUMMARY_1)) 

class TargetedPipelineModule(dspy.Module):
    def __init__(self, summarizer):
        super().__init__()
        # 只有这个是 dspy.Module 的属性，它们会被优化器调整
        self.summarizer = summarizer
                        
    def forward(self, dict_query, similar_threshold=DEFAULT_SIMILAR_THRESHOLD):
        # 调用 pipeline_func，使用优化目标和固定组件    
        res = pipeline_func(dict_query=dict_query,
                            embedding_dict_candidate=embedding_dict_candidate,                          
                            summarizer=self.summarizer,
                            similar_threshold=similar_threshold  # <-- 传入
                            )
        return dspy.Prediction(dict1=res['dict1'],
                               dict2=res['dict2'],
                               query2candidate=res['query2candidate'],
                               query_chains=res['query_chains'])
   
print('当前运行的单元格是In[5.2]，已经完成。')  

# In[5.3]使用最优的模型和final_threshold计算测试集指标值(OM)
# Calculate test set metrics using the optimal model and the final threshold(OM)

# =============================================================================
# 本单元格可以修改epoch_num_cv值，多次运行，用测试集计算交叉验证不同折保存模型的指标
# Modify epoch_num_cv and run multiple times to evaluate test set metrics for models saved across different CV folds
# =============================================================================

final_threshold={1:0.750946795,2:0.824800409,3:0.838501061,4:0.824121758,5:0.794613086}

print(f'当前测试集中数据共有{len(testset)}个。')
# num_selected=epoch_num_cv
num_selected=5
loaded_pipeline_2 = TargetedPipelineModule(summarizer=summarizer_candidate)
loaded_pipeline_2.load(".\\data\\dspy_opted_model\\opted_model\\optimized_dspy_False"+str(num_selected)+'.pkl')

scores_test_2 = [
    evalution_metric(example, loaded_pipeline_2(**example.inputs(), similar_threshold=final_threshold[num_selected]))
    for example in testset
    ]

# 过滤掉 -1
scores_test_2_filtered = [s for s in scores_test_2 if s != -1]

# 转为 numpy 数组计算平均值
scores_test_2_np = np.array(scores_test_2_filtered)
avg_scores_test_2 = np.mean(scores_test_2_np, axis=0)

print(f'***测试集在第{num_selected}折交叉验证优化后模型上的指标平均值：precision是{avg_scores_test_2[0]},recall是{avg_scores_test_2[1]},f1是{avg_scores_test_2[2]}。')

xlsx=openpyxl.load_workbook(r'.\\data\\result.xlsx')
sheet1=xlsx['test']
line=num_selected+1
sheet1['C'+str(line)]=avg_scores_test_2[0]
sheet1['D'+str(line)]=avg_scores_test_2[1]
sheet1['E'+str(line)]=avg_scores_test_2[2]
sheet1['F'+str(line)]=final_threshold[num_selected]   
sheet1['G'+str(line)]=len(testset) 
sheet1['H'+str(line)]=num_selected     
xlsx.save(r'.\\data\\result.xlsx')
xlsx.close()  # 确保文件被关闭
print('当前运行的单元格是In[5.3]，已经完成。')  

# In[5.4] 采用原始模型计算测试集得分(IM)
# Calculate test set metrics using the original model(IM)
#loaded_pipeline_1是优化前的原始模型
loaded_pipeline_1 = TargetedPipelineModule(summarizer=summarizer_1)

scores_test_1 = [
    evalution_metric(example, loaded_pipeline_1(**example.inputs(), similar_threshold=DEFAULT_SIMILAR_THRESHOLD))
    for example in testset
    ]

# 过滤掉 -1
scores_test_1_filtered = [s for s in scores_test_1 if s != -1]

# 转为 numpy 数组计算平均值
scores_test_1_np = np.array(scores_test_1_filtered)
avg_scores_test_1 = np.mean(scores_test_1_np, axis=0)

print(f'***测试集在原始模型上的指标平均值：precision是{avg_scores_test_1[0]},recall是{avg_scores_test_1[1]},f1是{avg_scores_test_1[2]}。')

xlsx=openpyxl.load_workbook(r'.\\data\\result.xlsx')
sheet1=xlsx['test']
line=7
sheet1['C'+str(line)]=avg_scores_test_1[0]
sheet1['D'+str(line)]=avg_scores_test_1[1]
sheet1['E'+str(line)]=avg_scores_test_1[2]
sheet1['F'+str(line)]=DEFAULT_SIMILAR_THRESHOLD
sheet1['G'+str(line)]=len(testset) 
sheet1['H'+str(line)]=num_selected   
xlsx.save(r'.\\data\\result.xlsx')
xlsx.close()  # 确保文件被关闭

print('当前运行的单元格是In[5.4]，已经完成。')  

# In[5.5] 提示词未优化，参数已经优化，计算测试集得分(TOM)
# Calculate test set metrics with an unoptimized prompt but optimized parameters(TOM)

#loaded_pipeline_1是优化前的原始模型
loaded_pipeline_1_1 = copy.deepcopy(loaded_pipeline_1)

scores_test_1_1 = [
    evalution_metric(example, loaded_pipeline_1_1(**example.inputs(), similar_threshold=final_threshold[num_selected]))
    for example in testset
    ]

# 过滤掉 -1
scores_test_1_1_filtered = [s for s in scores_test_1_1 if s != -1]

# 转为 numpy 数组计算平均值
scores_test_1_1_np = np.array(scores_test_1_1_filtered)
avg_scores_test_1_1 = np.mean(scores_test_1_1_np, axis=0)

print(f'***测试集在提示词未优化，参数已经优化的模型上的指标平均值：precision是{avg_scores_test_1_1[0]},recall是{avg_scores_test_1_1[1]},f1是{avg_scores_test_1_1[2]}。')

xlsx=openpyxl.load_workbook(r'.\\data\\result.xlsx')
sheet1=xlsx['test']
line=8
sheet1['C'+str(line)]=avg_scores_test_1_1[0]
sheet1['D'+str(line)]=avg_scores_test_1_1[1]
sheet1['E'+str(line)]=avg_scores_test_1_1[2]
sheet1['F'+str(line)]=final_threshold[num_selected]
sheet1['G'+str(line)]=len(testset) 
sheet1['H'+str(line)]=num_selected   
xlsx.save(r'.\\data\\result.xlsx')
xlsx.close()  # 确保文件被关闭

print('当前运行的单元格是In[5.5]，已经完成。')  

# In[5.6] 提示词已经优化，参数未优化，计算测试集得分(POM)
# Calculate test set metrics with an optimized prompt but unoptimized parameters(POM)

# loaded_pipeline_2 是第epoch_num_cv折交叉验证优化后的模型
loaded_pipeline_2_2=copy.deepcopy(loaded_pipeline_2)

scores_test_2_2 = [
    evalution_metric(example, loaded_pipeline_2_2(**example.inputs(), similar_threshold=DEFAULT_SIMILAR_THRESHOLD))
    for example in testset
    ]

# 过滤掉 -1
scores_test_2_2_filtered = [s for s in scores_test_2_2 if s != -1]

# 转为 numpy 数组计算平均值
scores_test_2_2_np = np.array(scores_test_2_2_filtered)
avg_scores_test_2_2 = np.mean(scores_test_2_2_np, axis=0)

print(f'***测试集在提示词已经优化，参数未优化的模型上的指标平均值：precision是{avg_scores_test_2_2[0]},recall是{avg_scores_test_2_2[1]},f1是{avg_scores_test_2_2[2]}。')

xlsx=openpyxl.load_workbook(r'.\\data\\result.xlsx')
sheet1=xlsx['test']
line=9
sheet1['C'+str(line)]=avg_scores_test_2_2[0]
sheet1['D'+str(line)]=avg_scores_test_2_2[1]
sheet1['E'+str(line)]=avg_scores_test_2_2[2]
sheet1['F'+str(line)]=DEFAULT_SIMILAR_THRESHOLD 
sheet1['G'+str(line)]=len(testset) 
sheet1['H'+str(line)]=num_selected        
xlsx.save(r'.\\data\\result.xlsx')
xlsx.close()  # 确保文件被关闭

print('当前运行的单元格是In[5.6]，已经完成。')  

