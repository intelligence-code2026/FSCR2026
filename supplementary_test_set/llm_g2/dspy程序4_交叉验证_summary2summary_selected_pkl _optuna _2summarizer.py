# In[0]进行基本设置，定义基本函数
import random
import time
import openpyxl
import dspy
import os
import re
import numpy as np
from openai import OpenAI
from sklearn.metrics.pairwise import cosine_similarity
import copy, pickle
import self_func
import sys
# os.environ["CUDA_VISIBLE_DEVICES"] = "0"  # 指定使用第0号GPU

num_cores = os.cpu_count()
print(f"CPU核心数: {num_cores}")

# =============================================================================
# epoch_num_cv是交叉验证轮数，可以修改！！！！！！！！！！！！！！！！！！！！！！！！！！
# =============================================================================
epoch_num_cv=1      #这个值最大是5，每轮验证完要修改这个值
assert 1 <= epoch_num_cv <= 5, "epoch_num_cv 应为 1 到 5 的整数"

# =============================================================================
# num_threads_lim是多线程计算的线程上限，可以修改！！！！！！！！！！！！！！！！！！
# =============================================================================
num_threads_lim = 8

# =============================================================================
# num_query是生成的查询案例数量，包括训练集和验证集，可以修改！！！！！！！！！！！！！！！
# =============================================================================
num_query = 800

# =============================================================================
# num_testset是生成的测试集中数据量，可以修改！！！！！！！！！！！！！！！！！！！！！！！
# =============================================================================
num_testset=200

# =============================================================================
# num_candidate_case是是生成查询案例时，引用的候选案例数量，可以修改！！！！！！！！
# =============================================================================
num_candidate_case=3

# =============================================================================
# num_summary是生成查询案例时，每个候选案例引用的summary数量，可以修改！！！！！！！
# =============================================================================
num_summary = 2

# =============================================================================
#num_candidate是候选案例数量，可以修改！！！！！！！！！！！！！！！！！！！！！！！
# =============================================================================
num_candidate = 37

# =============================================================================
# max_tokens_lm是大模型输出的tokens限值，可以修改！！！！！！！！！！！！！！！
# =============================================================================
max_tokens_lm = None

# =============================================================================
# similar_threshold_value是细粒度类案检索相似的阈值，可以修改！！！！！！！！！！！
# =============================================================================
# similar_threshold_value = 0.74    	      #目前similar_threshold_value=0.89为最优
DEFAULT_SIMILAR_THRESHOLD = 0.9

# =============================================================================
# num_candidate_selected_value是在build_dict2_from_query_chains函数中，
# 根据与query的相似度筛选出candidate的数量，可以修改！！！！！！！！！！！！！！！！
# =============================================================================
num_candidate_selected_value=37

# =============================================================================
# auto_option是优化参数设置，可以选择None、"light"、"medium"、"heavy"！！！！！！！
# =============================================================================
auto_option = "heavy"

# 并发线程数（实际用于 ThreadPoolExecutor）
num_threads = min(num_threads_lim, num_cores)

# 读取 API Key
try:
    xlsx_api_key = openpyxl.load_workbook(r'..\data\api_key.xlsx')
    sheet_api_key = xlsx_api_key['Sheet1']
    api_key_wenxin = sheet_api_key['B4'].value
    api_key_kimi = sheet_api_key['B' + str(6)].value
except Exception as e:
    print('未找到llm的api_key！！！')
    print(f"错误详情: {e}")
    sys.exit(1)

try:             
    dict_case_text_pickle=r'.\data\dspy_opted_model\dict_case_text.pkl'
    with open(dict_case_text_pickle,'rb') as f_obj9:
         dict_case_text=pickle.load(f_obj9)     
    
except FileNotFoundError:    
    print('加载dict_case_text时出错未找到该文件，要重新生成。')
    # 读取候选案例文本
    xlsx_candidation = openpyxl.load_workbook(r'.\data\候选案例2.xlsx')
    sheet_candidate = xlsx_candidation['Sheet2']
    total_rows = sheet_candidate.max_row
    dict_case_text = {}
    for k in list(range(2, total_rows + 1))[:num_candidate]:
        dict_case_text[k] = sheet_candidate['B' + str(k)].value
    xlsx_candidation.close()
                 

    dict_case_text_pickle = r'.\data\dspy_opted_model\dict_case_text.pkl'
    with open(dict_case_text_pickle,'wb') as f_obj8:
          pickle.dump(dict_case_text, f_obj8)      

print(f'由xlsx生成的dict_case_text包含的数据共{len(dict_case_text)}条。')



# 配置 LLM

# lm = dspy.LM("openai/qwen-turbo", api_key=api_key, max_tokens=max_tokens_lm,api_base="https://dashscope.aliyuncs.com/compatible-mode/v1", temperature=0.5)

# lm_high_temperature = dspy.LM("openai/qwen-turbo", api_key=api_key, max_tokens=max_tokens_lm,api_base="https://dashscope.aliyuncs.com/compatible-mode/v1", temperature=1.0)

# deepseek_lm = dspy.LM("openai/deepseek-chat", api_key=api_key_deepseek, max_tokens=max_tokens_lm,api_base="https://api.deepseek.com",temperature=0.2)


# deepseek_lm = dspy.LM("openai/qwen-turbo", api_key=api_key, max_tokens=max_tokens_lm,api_base="https://dashscope.aliyuncs.com/compatible-mode/v1", temperature=0.1)

# lm_high_temperature = dspy.LM("openai/deepseek-chat", api_key=api_key_deepseek, max_tokens=max_tokens_lm,api_base="https://api.deepseek.com",temperature=0.3)

# lm = dspy.LM("openai/deepseek-chat", api_key=api_key_deepseek, max_tokens=max_tokens_lm,api_base="https://api.deepseek.com",temperature=0.2)

lm = dspy.LM("openai/ernie-4.5-turbo-128k", api_key=api_key_wenxin, max_tokens=max_tokens_lm,api_base="https://qianfan.baidubce.com/v2", temperature=0.1)

lm_high_temperature = dspy.LM("openai/ernie-4.5-turbo-128k", api_key=api_key_wenxin, max_tokens=max_tokens_lm,api_base="https://qianfan.baidubce.com/v2",temperature=0.5)

deepseek_lm = dspy.LM("openai/kimi-latest", api_key=api_key_kimi, max_tokens=max_tokens_lm,api_base="https://api.moonshot.cn/v1",temperature=0.2)


dspy.configure(lm=lm)

# clien是词嵌入大模型
client = OpenAI(
   api_key=sheet_api_key['B'+str(2)].value,  
   base_url="https://dashscope.aliyuncs.com/compatible-mode/v1"  )

# ========================================
# 初始提示词（会被优化）
# ========================================
# file_path_summary = r'.\data\prompt\detail_summary_prompt.txt'    #读取提示词文本文件

# =============================================================================
# PROMPT_SUMMARY是用于抽取查询案例摘要的提示词
# =============================================================================
#优化后的summary提示词，每次优化完成后都会更新，在优化单元格有更新语句
# file_path_summary = r'.\data\prompt\detail_summary_prompt_opt'+str(epoch_num_cv)+'.txt'    #读取优化后的提示词文本文件
file_path_summary = r'..\data\prompt\detail_summary_prompt_opt0-2.txt'    #读取原始的提示词文本文件
with open(file_path_summary, 'r', encoding='utf-8') as file:       
      detail_summary_prompt = file.read()      
PROMPT_SUMMARY = detail_summary_prompt      
# PROMPT_SUMMARY = "提示词1：你是一个法律专家，请用一句话总结下面的文本要点，可以输出多条摘要，每条摘要为一句话，多条请用#分割，不用换行。输出必须为中文。"


# =============================================================================
# PROMPT_SUMMARY_CANDIDATE是用于抽取候选案例摘要的提示词
# =============================================================================
file_path_summary_candidate = r'..\data\prompt\detail_summary_prompt_opt0-3.txt'    #读取原始的提示词文本文件
with open(file_path_summary_candidate, 'r', encoding='utf-8') as file_summary_candidate:       
      detail_summary_prompt_candidate = file_summary_candidate.read()      
PROMPT_SUMMARY_CANDIDATE = detail_summary_prompt_candidate     


# file_path_kg = r'.\data\prompt\judgement_rule_kg.txt'    #读取提示词文本文件
# # file_path_kg = r'.\data\prompt\candidate_summary_factor_prompt.txt'     #读取提示词文本文件
# with open(file_path_kg, 'r', encoding='utf-8') as file_kg:       
#      kg_prompt = file_kg.read()
# PROMPT_ACTION_CHAIN = kg_prompt
PROMPT_ACTION_CHAIN = "提示词2：请从下列摘要中提取关键行为链（用'->'连接每个动作），每行对应一条摘要的行为链，若一条摘要有多条行为链用;分开。输出必须为中文。"


# file_path_compose = r'.\data\prompt\compose_prompt.txt'    #读取提示词文本文件
file_path_compose = r'..\data\prompt\compose_prompt - 准确率高.txt'    #读取提示词文本文件
with open(file_path_compose, 'r', encoding='utf-8') as file_compose:       
      compose_prompt = file_compose.read()
PROMPT_COMPOSE_QUERY = compose_prompt     
# PROMPT_COMPOSE_QUERY = "提示词3：请根据以下行为链（用;分隔各链）编写一个完整的不超过3000个字的法律案例，要求包含所有行为链且前后逻辑自洽，可以修改人名、地名、单位名称等。"
# PROMPT_COMPOSE_QUERY = '以下文本包括了若干个法律事实，你根据这些法律事实另外编写一个完整的建设工程合同纠纷的案例，要求是一个完整案件，可以重新编写涉案主体的名称，不能是多个案件的集合。你编写的文字仅仅包括案情描述，包括日期，地点等细节的描述要比较准确，字数1000字左右。'   


# file_path_check = r'.\data\prompt\check_prompt.txt'    #读取提示词文本文件
# with open(file_path_check, 'r', encoding='utf-8') as file_check:       
#      check_prompt = file_check.read()
# PROMPT_JUDGE_QUERY = check_prompt     
# PROMPT_JUDGE_QUERY = "提示词4：命名为chains的文本是若干要点（用;分隔各要点），命名为query的文本是根据chains生成的文字，检查一下query中是否包含了大部分给定的要点，并判断文本逻辑是否一致。满足条件，输出yes，不满足条件，输出no并给出理由，理由不超过50字。"
PROMPT_JUDGE_QUERY = '以下文本中，命名为chains的文字是若干要点（用;分隔各要点），命名为query的文字是一个法律案例，你检查一下query是否具备以下要求：1.案情完整\n 2.逻辑无错误\n 3.包括了chain的全部情节，当事人名称、地点名称、时间可与chain不同。\n如果具备上述的全部要求，则输出标志语：yes，如果不完全具备上述的全部要求，输出标志语：no，标志语后面输出判断理由，标志语和判断理由之间用,分割。'   


# ========================================
# 定义 Signature
# ========================================
# def make_signature(input_name: str, output_name: str, prompt: str):
#     Sig = type(
#         "Sig",
#         (dspy.Signature,),
#         {
#             input_name: dspy.InputField(),
#             output_name: dspy.OutputField(),
#         }
#     )
#     Sig.__doc__ = prompt
#     return Sig

def make_signature(input_name: str, output_name: str, prompt: str):
    Sig = type(
        "Sig",
        (dspy.Signature,),
        {
            "__annotations__": {input_name: str, output_name: str},
            input_name: dspy.InputField(),
            output_name: dspy.OutputField(),
        }
    )
    Sig.__doc__ = prompt
    return Sig


summarizer = dspy.Predict(make_signature("text", "summary", PROMPT_SUMMARY))

summarizer_candidate = dspy.Predict(make_signature("text", "summary", PROMPT_SUMMARY_CANDIDATE))

action_chain_extractor = dspy.Predict(make_signature("summary", "action_chain", PROMPT_ACTION_CHAIN))
composer = dspy.Predict(make_signature("action_chains", "composed_text", PROMPT_COMPOSE_QUERY))
judge = dspy.Predict(make_signature("composed_text", "judge_result", PROMPT_JUDGE_QUERY))

# ========================================
# 工具函数 (MODIFIED: 接收 Predictor 参数)
# ========================================
def safe_split(s:str):
    # 步骤1：合并多个连续的 # 为一个
    s_clean = re.sub(r'#{2,}', '#', s)
    
    # 步骤2：按 # 分割，并去除空白项（可选）
    result = [x.strip() for x in s_clean.split('#') if x.strip()]
    
    return result

def call_summarizer(summarizer, text: str)->dict:
    dict_summary_temp={}
    p = summarizer(text=text) # **使用传入的参数**
    summary_text = getattr(p, "summary", "")
    try:
        list_summary=safe_split(summary_text)
    except:
        list_summary=[summary_text]
    for i in range(len(list_summary)):
        dict_summary_temp[i]=list_summary[i]
    return dict_summary_temp

def call_summarizer_candidate(summarizer_candidate, text: str)->dict:
    dict_summary_temp={}
    p = summarizer_candidate(text=text) # **使用传入的参数**
    summary_text = getattr(p, "summary", "")
    try:
        list_summary=safe_split(summary_text)
    except:
        list_summary=[summary_text]
    for i in range(len(list_summary)):
        dict_summary_temp[i]=list_summary[i]
    return dict_summary_temp

def call_action_extractor(action_chain_extractor, dict_summary_temp: dict)->dict:
    dict_chain_temp={}
    for j in dict_summary_temp.keys():
        try:
            p = action_chain_extractor(summary=dict_summary_temp[j]) # **使用传入的参数**
            chain_gen = getattr(p, "action_chain", "")
            # if isinstance(chain_gen, str) and chain_gen.strip():
            if chain_gen and chain_gen.strip():  
                dict_chain_temp[j]=chain_gen
            else:
                continue
        except Exception as e:
            continue       
    return dict_chain_temp

def call_composer(composer, dict_chain:dict)->str:
    num_factor=0
    input_text=''
    for key1 in dict_chain.keys():
        for key2 in dict_chain[key1].keys():
            num_factor+=1
            factor_text=dict_chain[key1][key2]
            input_text+=f'{num_factor}.{factor_text};'
    with dspy.context(lm=lm_high_temperature):
        p = composer(action_chains=input_text) 
    return getattr(p, "composed_text", ""), input_text

def call_judge(judge, query: str, chains:str, use_deepseek=True)->str:
    # 使用 deepseek_lm 检查 query 是否包含所有行为链且逻辑一致
    input_text=f'chains是{chains}。query是{query}'
    if use_deepseek:
        with dspy.context(lm=deepseek_lm):
            p = judge(composed_text=input_text) # **使用传入的参数**
    else:
        p = judge(composed_text=input_text) # **使用传入的参数**
    return getattr(p, "judge_result", "")

def embedding_vector(sentence:str,client=client)->list:     
    completion = client.embeddings.create(
        model="text-embedding-v3",
        input=sentence,
        dimensions=1024,
        encoding_format="float"
    )
    return  completion.data[0].embedding

def similiar_score(embedding_vec_query, embedding_vec_candidate)->float:
    # 确保输入是numpy数组（如果你的输入已经是数组，可以省略这步）
    vec1 = np.array(embedding_vec_query).reshape(1, -1)  # 强制转为2D: (1, n)
    vec2 = np.array(embedding_vec_candidate).reshape(1, -1)  # 强制转为2D: (1, n)
    
    # 计算余弦相似度（结果是一个1x1矩阵）
    similarity_matrix = cosine_similarity(vec1, vec2)
    
    # 返回标量值（提取矩阵中的单个数值）
    return similarity_matrix[0][0]

def generate_query_and_validate(composer, judge, dict_chain, dict1, max_retries=5)->dict:
    """
    用 lm 生成 query，并用 deepseek_lm 判定是否包含全部链且逻辑一致；
    若不符合则重试。最多尝试 max_retries + 1 次（含第一次）。
    """
    for attempt in range(max_retries + 1):
        try:
            query_text,chains = call_composer(composer,dict_chain)
            # print(f'***生成的query是:{query_text}')
            # print(f'***生成的chains是:{chains}')
            
            result_check = call_judge(judge, query_text, chains, use_deepseek=True)
            
            # result_check='yes'
            
            
            print(f'***对query检查结果是:{result_check}')
            if 'yes' in result_check.strip().lower():
                return {'text': query_text, 'label': dict1}
        except Exception as e:
            print(f"第 {attempt + 1} 次调用失败: {e}")
        
        # 如果不是最后一次尝试，等待后继续
        if attempt < max_retries:
            time.sleep(2)    
    # 所有尝试都失败
    print(f'共进行了 {max_retries + 1} 次查询生成，均失败。')
    return {'text': 'failed', 'label': dict1}

# 对如下函数进行了修改，返回的是抽取的摘要字典*********************************************************
# def process_query_extract(summarizer, action_chain_extractor, dict_query)->dict:
def process_query_extract(summarizer, dict_query)->dict:    
    """对 query 用提示词1和提示词2再次抽取摘要与行为链"""
    try:
        dict_summary_temp = call_summarizer(summarizer, dict_query['text']) 
        return {'chain':dict_summary_temp,'label':dict_query['label']}
    except Exception as e:        
        print(f"生成query的dict_summary时，调用失败: {e}")
        return 'gen_query_summary_failed'   
    
    # try:
    #     dict_chain_temp = call_action_extractor(action_chain_extractor, dict_summary_temp)
    # except Exception as e:            
    #     print(f"生成query的dict_chain_temp时，调用失败: {e}") 
    #     return 'gen_query_chain_failed'    
    # return {'chain':dict_chain_temp,'label':dict_query['label']}

# 对如下函数进行了修改，返回的是抽取的摘要字典************************************************************
def process_candidate_extract(summarizer_candidate, action_chain_extractor, dict_case_text):
    """对 candidate 用提示词1和提示词2抽取摘要与行为链，并获得词向量"""
    dict_candidate_summary={}
    dict_candidate_chain={}
    dict_candidate_embedding={}
    for key_candidate in dict_case_text.keys():
        try:
            dict_summary_candidate_temp = call_summarizer_candidate(summarizer_candidate, dict_case_text[key_candidate])
            dict_candidate_summary[key_candidate]=dict_summary_candidate_temp
        except Exception as e:        
            print(f"生成第{key_candidate}个候选案例的摘要时，调用失败: {e}")
            continue
    for key_candidate_summary in dict_candidate_summary.keys():
        sub_dict_candidate_chain={}
        sub_dict_candidate_embedding={}
        for key_candidate_summary_sub in dict_candidate_summary[key_candidate_summary].keys():
            try:
                chain_extracted_candidate=action_chain_extractor(
                    summary=dict_candidate_summary[key_candidate_summary][key_candidate_summary_sub])   #获取某条summary的行为链
                chain_gen_candidate = getattr(chain_extracted_candidate, "action_chain", "")
                # if isinstance(chain_gen_candidate,str) and chain_gen_candidate.strip(): 
                if chain_gen_candidate and chain_gen_candidate.strip():          
                    try:
                        # 对如下语句进行了修改*******************************************************
                        # sub_dict_candidate_embedding[key_candidate_summary_sub]=embedding_vector(chain_gen_candidate)
                        sub_dict_candidate_embedding[key_candidate_summary_sub]=embedding_vector(dict_candidate_summary[key_candidate_summary][key_candidate_summary_sub])
                        sub_dict_candidate_chain[key_candidate_summary_sub]=chain_gen_candidate
                    except Exception as e:
                        print(f'第{key_candidate_summary}个候选案例的第{key_candidate_summary_sub}摘要的行为链词向量生成失败：{e}')
                        continue
                else:
                    continue
            except Exception as e: 
                print(f"生成第{key_candidate_summary}个候选案例的第{key_candidate_summary_sub}个行为链时，调用失败: {e}")
                continue
        dict_candidate_chain[key_candidate_summary]=sub_dict_candidate_chain
        dict_candidate_embedding[key_candidate_summary]=sub_dict_candidate_embedding
    
    return dict_candidate_summary,dict_candidate_chain,dict_candidate_embedding
       

def build_dict2_from_query_chains(embedding_dict_query, 
                                  embedding_dict_candidate_temp,
                                  num_candidate_selected=num_candidate_selected_value ,
                                  similar_threshold=DEFAULT_SIMILAR_THRESHOLD)->dict:
    
    list_key_candidate=self_func.rank_by_bidirectional_similarity(embedding_dict_query, embedding_dict_candidate_temp,num_candidate_selected)    
    embedding_dict_candidate={}
    for k in list_key_candidate:
        embedding_dict_candidate[k]=embedding_dict_candidate_temp[k]
    print(f'embedding_dict_candidate中有{len(embedding_dict_candidate)}个字典。')
    
    for key_query in embedding_dict_query.keys():    
        similar_dict_key_query={}    
        for q in embedding_dict_query[key_query].keys():
            similar_dict_key_candidate={}
            # for k in embedding_dict_candidate.keys():    # 加载全部候选案例
            for k in embedding_dict_candidate.keys():      # 仅加载用于生成查询案例的候选案例 
                similar_dict_candidate_k={}         
                for c in embedding_dict_candidate[k].keys():
                    if isinstance(embedding_dict_candidate[k][c], list) and isinstance(embedding_dict_query[key_query][q], list):
                        similiar_value=similiar_score(embedding_dict_query[key_query][q],embedding_dict_candidate[k][c])
                    else :
                        similiar_value=0
                    similar_dict_candidate_k[c]=copy.deepcopy(similiar_value)                    
                similar_dict_key_candidate[k]= copy.deepcopy(similar_dict_candidate_k)      
            similar_dict_key_query[q]= copy.deepcopy(similar_dict_key_candidate)  
                 
        # =============================================================================
        # 以下程序可以根据similar_threshold值求得与查询案例各条摘要具有最大相似度的查询案例摘要的编号
        # =============================================================================
        # similar_threshold=0.877 #可以修改相似度筛选阈值！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！
        max_similar_dict_key_query={}                
        for q2 in similar_dict_key_query.keys():
            max_condidate_of_condidate_case={}
            
            for k2 in similar_dict_key_query[q2].keys():
                condidate_of_condidate_case={}
               
                for c2 in similar_dict_key_query[q2][k2].keys():           
                    if similar_dict_key_query[q2][k2][c2]>similar_threshold:
                        condidate_of_condidate_case[c2]=similar_dict_key_query[q2][k2][c2]
                if condidate_of_condidate_case!={}:       
                    max_key_condidate = max(condidate_of_condidate_case, key= condidate_of_condidate_case.get)   #取得具有最大相似度的键  
                    
                    #max_condidate_of_condidate_case是一个字典，键为第k2个候选案例，值为与查询案例第q2条案情摘要具有最大相似度的候选案例的摘要编号    
                    max_condidate_of_condidate_case[k2]=max_key_condidate
                                                
                else:
                #     max_key_condidate=None
                    continue
                
            if max_condidate_of_condidate_case!={}:  
                # print(len(max_condidate_of_condidate_case))
                if len(max_condidate_of_condidate_case)>1:
                    max_similar_dict={}
                    similiar_dict={}
                    for key in max_condidate_of_condidate_case.keys():
                        similiar_dict[key]=similar_dict_key_query[q2][key][max_condidate_of_condidate_case[key]]               
                    max_key = max(similiar_dict, key=similiar_dict.get)  # 找出最大值对应的键
                    max_similar_dict[max_key]=max_condidate_of_condidate_case[max_key]
                elif len(max_condidate_of_condidate_case)==1:
                    max_similar_dict=max_condidate_of_condidate_case
                # print(f'max_similar_dict的数据个数是{len(max_similar_dict)}')
                max_similar_dict_key_query[q2]=max_similar_dict  
            else:
                continue                        
        query_pre_dict = {}
        
        # 遍历原始字典a
        for outer_key, inner_dict in max_similar_dict_key_query.items():
            for inner_key, value in inner_dict.items():
                if inner_key not in query_pre_dict:
                    query_pre_dict[inner_key] = set()  # 使用集合自动去重
                query_pre_dict[inner_key].add(value)  # 添加值到集合
        
        # 将集合转换回列表（可选，如果希望结果是列表而不是集合）
        dict2 = {k: list(v) for k, v in query_pre_dict.items()}
        print(dict2)
    return dict2


def build_dict2_from_query_chains_2(embedding_dict_query, 
                                    embedding_dict_candidate_temp,
                                    num_candidate_selected=num_candidate_selected_value ,
                                    similar_threshold=DEFAULT_SIMILAR_THRESHOLD)-> tuple[dict, dict]:
    
    '''输出的output键是embedding_dict_query[0]对应的键，值是一个元组，
    该元组第一个元素是 embedding_dict_candidate_temp的键（即第几个候选案例）
    该元组第二个元素是embedding_dict_candidate_temp中作为值的字典的键（即某个候选案例的第几个摘要或者链）.
    输出的dict2是一个字典，是对dict1的预测值。'''
    
    list_key_candidate=self_func.rank_by_bidirectional_similarity(embedding_dict_query, embedding_dict_candidate_temp,num_candidate_selected)    
    embedding_dict_candidate={}
    for k in list_key_candidate:
        embedding_dict_candidate[k]=embedding_dict_candidate_temp[k]
    print(f'embedding_dict_candidate中有{len(embedding_dict_candidate)}个字典。')
        
    num_candidate_embedding=0
    list_candidate_embedding=[]
    dict_site={}
    for k1 in embedding_dict_candidate.keys():        
        for k2 in embedding_dict_candidate[k1].keys():
            list_candidate_embedding.append(embedding_dict_candidate[k1][k2])
            dict_site[num_candidate_embedding]=(k1,k2)
            num_candidate_embedding+=1
    
    k5=0
    site_query={}
    for k4 in embedding_dict_query[0].keys():
        site_query[k5]=k4
        k5+=1
        
    array_candidate_embedding=np.array(list_candidate_embedding)
    embedding_dict_query_temp = list(embedding_dict_query.values())[0]
    array_query_embedding = np.array(list(embedding_dict_query_temp.values()))
    qc_sim_embedding = cosine_similarity(array_query_embedding, array_candidate_embedding)
    print(f'embedding_dict_query共有{len(embedding_dict_query[0])}个词向量，qc_sim_embedding的形状是{qc_sim_embedding.shape}')
    mask = qc_sim_embedding > similar_threshold
        
    # 如果某行没有满足条件的元素，则该行最大值设为 -inf，避免干扰 argmax
    qc_masked = np.where(mask, qc_sim_embedding, -np.inf)

    # 每行第一个最大值的列索引（argmax 默认返回第一个最大值的位置）
    col_indices = np.argmax(qc_masked, axis=1)

    # 判断每行是否至少有一个元素 > threshold
    valid_rows = np.any(mask, axis=1)

    # 获取有效的行索引
    row_indices = np.where(valid_rows)[0]
    col_indices = col_indices[valid_rows]

    # 转换为 (行, 列) 元组的列表,result是一个列表，列表每个元素都是一个元组，
    # 元组第一个值是选中的qc_sim_embedding中元素的行号，元组第二个值是列号
    result = [(int(row), int(col)) for row, col in zip(row_indices, col_indices)]
    output={}
    for k3 in range(len(result)):
        output[site_query[result[k3][0]]]= dict_site[result[k3][1]]
        
    dict2 = {}
    for x, y in output.values():
        dict2.setdefault(x, []).append(y)
   
    return output,dict2


# ========================================
# pipeline 函数 (MODIFIED: 传递 Predictor 实例)
# ========================================
def pipeline_func(dict_query, 
                  embedding_dict_candidate, 
                  summarizer,
                  similar_threshold=DEFAULT_SIMILAR_THRESHOLD
                  ):
                   
    #  对 query 抽取摘要和行为链
    embedding_dict_query_temp={}
    if dict_query['text']!='failed':
        query_chains = process_query_extract(summarizer, dict_query) # **传递 Predictor**
        if isinstance(query_chains, dict):
            for key_query_chains in query_chains['chain'].keys():
                if query_chains['chain'][key_query_chains]:
                    embedding_dict_query_temp[key_query_chains]=embedding_vector(query_chains['chain'][key_query_chains])
                else:
                    continue
        else:
            print(f'query_chain不是个dict，而是{query_chains},出错了！！！！！')
            return {"dict1": {},"dict2": {}}
    else:
        print(f'当前的dict_query是{dict_query},出错了！！！！！')
        return {"dict1": {},"dict2": {}}
    
    embedding_dict_query={0:embedding_dict_query_temp}                              
        
    try:
        if embedding_dict_query[0]:
            query2candidate,dict2 = build_dict2_from_query_chains_2(
                                    embedding_dict_query, 
                                    embedding_dict_candidate,
                                    similar_threshold=similar_threshold  # <-- 传入阈值
                                                                    )
        else:
            print(f'当前的embedding_dict_query[0]是{embedding_dict_query[0]},出错了！！！！！')
            dict2 = {}
            query2candidate={}
    except Exception as e:
        print(f"构建dict2失败: {e}")
        dict2 = {}
        query2candidate={}
    return {"dict1": dict_query['label'],"dict2": dict2, "query2candidate":query2candidate,"query_chains":query_chains }

def my_metric(example, pred, trace=None):
    
    if pred.dict2=={} or pred.dict1=={}:
        
        return -1
    else:
        precision,recall,f1=self_func.metrics_func(pred.dict2, pred.dict1)     
        return float(f1)
    
def evalution_metric(example, pred, trace=None):
    if pred.dict2=={} or pred.dict1=={}:
        return -1
    else:
        precision,recall,f1=self_func.metrics_func(pred.dict2, pred.dict1)     
        return [precision,recall,f1]    
print('当前运行的单元格是In[0]，已经完成。')    
# In[1]生成candidate，query，训练集，验证集
try:
    dict_candidate_summary_pickle=r'.\data\dspy_opted_model\dict_candidate_summary.pkl'
    with open(dict_candidate_summary_pickle,'rb') as f_obj1:
         dict_candidate_summary=pickle.load(f_obj1) 
         
    dict_candidate_chain_pickle=r'.\data\dspy_opted_model\dict_candidate_chain.pkl'
    with open(dict_candidate_chain_pickle,'rb') as f_obj2:
         dict_candidate_chain=pickle.load(f_obj2) 
         
    embedding_dict_candidate_pickle=r'.\data\dspy_opted_model\embedding_dict_candidate.pkl'
    with open(embedding_dict_candidate_pickle,'rb') as f_obj3:
         embedding_dict_candidate=pickle.load(f_obj3)              
except FileNotFoundError:
    dict_candidate_summary,dict_candidate_chain,embedding_dict_candidate=process_candidate_extract(
        summarizer_candidate, 
        action_chain_extractor, 
        dict_case_text)
    
    dict_candidate_summary_pickle=r'.\data\dspy_opted_model\dict_candidate_summary.pkl'  
    with open(dict_candidate_summary_pickle,'wb') as f_obj4:
         pickle.dump(dict_candidate_summary, f_obj4)
    
    dict_candidate_chain_pickle=r'.\data\dspy_opted_model\dict_candidate_chain.pkl'  
    with open(dict_candidate_chain_pickle,'wb') as f_obj5:
         pickle.dump(dict_candidate_chain, f_obj5)  
         
    embedding_dict_candidate_pickle=r'.\data\dspy_opted_model\embedding_dict_candidate.pkl'  
    with open(embedding_dict_candidate_pickle,'wb') as f_obj6:
         pickle.dump(embedding_dict_candidate, f_obj6)   
         
try:
    list_query_pickle=r'.\data\dspy_opted_model\list_query.pkl'
    with open(list_query_pickle,'rb') as f_obj10:
         list_query=pickle.load(f_obj10)    
    assert len(list_query) == num_query,f"加载的查询案例 {len(list_query)} 与预期 {num_query} 不符"
except (FileNotFoundError, EOFError, pickle.UnpicklingError, AssertionError) as e_list_query:    
    print('加载list_query时出错',e_list_query)

    if isinstance(e_list_query, FileNotFoundError):
        print('list_query文件未找到，需要重新生成。')
        list_query=[]
    elif isinstance(e_list_query, AssertionError):
        print(f'list_query加载的查询案例不够，需要补充生成{num_query-len(list_query)}条案例数据。')
        
    add_num_query=num_query-len(list_query)
    for _ in range(add_num_query): 
         dict1={}
         dict_chain={}
         list_candidate=random.sample(list(embedding_dict_candidate.keys()), num_candidate_case)  
         for k_dict in list_candidate:
# =============================================================================
#         #从每个候选案例的摘要字典中随机抽取键构建dict1
# =============================================================================
        # dict_summaries = dict_candidate_summary[key_dict]
        # list_summary_label=random.sample(list(dict_summaries.keys()), min(num_summary,len(dict_summaries.keys())))
        
# =============================================================================
#         #从每个候选案例的行为链字典中随机抽取键构建dict1
# =============================================================================
        # dict_summaries = dict_candidate_chain[key_dict]
        # list_summary_label=random.sample(list(dict_summaries.keys()), min(num_summary,len(dict_summaries.keys())))

# =============================================================================
#         #从每个候选案例的行为链词嵌入字典中随机抽取键构建dict1
# =============================================================================
             dict_summaries = embedding_dict_candidate[k_dict]
             list_summary_label=random.sample(list(dict_summaries.keys()), min(num_summary,len(dict_summaries.keys())))  
        
             dict1[k_dict]=copy.deepcopy(list_summary_label)
             sub_dict_chain={}
             for key_list in list_summary_label:
          
             # chain_extracted=action_chain_extractor(summary=dict_summaries[key_list])   #获取某条summary的行为链
             # chain_gen = getattr(chain_extracted, "action_chain", "")
             # sub_dict_chain[key_list]=chain_gen
                 try:
#对如下语句进行修改，用summary代替chain生成query ****************************************************            
                     # sub_dict_chain[key_list]=dict_candidate_chain[k_dict][key_list]
                     sub_dict_chain[key_list]=dict_candidate_summary[k_dict][key_list]
                 except Exception as e:
                     print(f'在第{k_dict}个候选案例的第{key_list}摘要调用对应行为链出错',e)
                     print(f'第{k_dict}个候选案例的摘要键列表为{list_summary_label}，行为链键列表为{list(dict_candidate_chain[k_dict])}')
             dict_chain[k_dict]=sub_dict_chain            
         dict_query=generate_query_and_validate(composer, judge, dict_chain, dict1, max_retries=5)
         if dict_query['text']=='failed':
              continue
         list_query.append(dict_query)
         
    list_query_pickle=r'.\data\dspy_opted_model\list_query.pkl'  
    with open(list_query_pickle,'wb') as f_obj11:
         pickle.dump(list_query, f_obj11)  

dataset = []                             #dataset是全部数据的集合
for query_dict in list_query:   
    example = dspy.Example(dict_query=query_dict, label=query_dict['label']).with_inputs("dict_query")
    dataset.append(example)
           
step = len(dataset) // 5  #  //是整除运算符号
start = step * (epoch_num_cv - 1)
end = step * epoch_num_cv if epoch_num_cv != 5 else len(dataset)

val_set = dataset[start:end]  #val_set是验证集合
trainset = dataset[:start] + dataset[end:]   #trainset是训练集合
print(f'dataset中共有数据{len(dataset)}条。')
print(f'trainset包含的数据共{len(trainset)}条。')
print(f'val_set包含的数据共{len(val_set)}条。')
print('当前运行的单元格是In[1]，已经完成。')
# In[2]进行优化
start_time=time.time()
# 使用 dspy 的优化器
# optimizer = dspy.MIPROv2(metric=my_metric, auto="light")
# optimizer = dspy.MIPROv2(metric=my_metric, auto="light",num_threads=num_cores)
optimizer = dspy.MIPROv2(metric=my_metric, 
                         # num_candidates=20,  #num_candidates常设为10~50，数据量大或任务复杂时可适当增大
                         auto=auto_option,                       
                         num_threads = min(num_threads_lim, num_cores),
                         max_bootstrapped_demos=6,  #每次优化最多自动生成的few-shot示例数量
                         max_labeled_demos=6, # 每次优化最多从trainset中采样的带标签示例数量                        
                         metric_threshold=0.6,  # 设定合格阈值
                         verbose=True
                         )

print("\n开始调用 dspy 优化器（仅优化 PROMPT_SUMMARY 和 PROMPT_ACTION_CHAIN）...")

# 创建一个新的 PipelineModule，只将需要优化的组件注册为 dspy.Module 属性
class TargetedPipelineModule(dspy.Module):
    def __init__(self, summarizer):
        super().__init__()
        # 只有这个是 dspy.Module 的属性，它们会被优化器调整
        self.summarizer = summarizer
        # self.action_chain_extractor = action_chain_extractor
                        
    def forward(self, dict_query, similar_threshold=DEFAULT_SIMILAR_THRESHOLD):
        # 调用 pipeline_func，使用优化目标和固定组件    
        res = pipeline_func(dict_query=dict_query,
                            embedding_dict_candidate=embedding_dict_candidate,                          
                            summarizer=self.summarizer,
                            similar_threshold=similar_threshold  # <-- 传入
                            )
        return dspy.Prediction(dict1=res['dict1'],
                               dict2=res['dict2'],
                               query2candidate=res['query2candidate'],
                               query_chains=res['query_chains'])
  
# TargetedPipelineModule实例化
pipeline_module_targeted = TargetedPipelineModule(summarizer=summarizer)

# 优化前评估（使用默认阈值）

scores_before = [
    evalution_metric(example, pipeline_module_targeted(**example.inputs(), similar_threshold=DEFAULT_SIMILAR_THRESHOLD))
    for example in val_set
    ]

# 过滤掉 -1
scores_before_filtered = [s for s in scores_before if s != -1]

# 转为 numpy 数组计算平均值
scores_before_np = np.array(scores_before_filtered)
avg_scores_before = np.mean(scores_before_np, axis=0)

xlsx=openpyxl.load_workbook(r'.\\data\\result.xlsx')
sheet1=xlsx['CV']
line=(epoch_num_cv-1)*3+2
sheet1['C'+str(line)]=avg_scores_before[0]
sheet1['D'+str(line)]=avg_scores_before[1]
sheet1['E'+str(line)]=avg_scores_before[2]
sheet1['G'+str(line)]=DEFAULT_SIMILAR_THRESHOLD
sheet1['H'+str(line)]=len(trainset)
sheet1['I'+str(line)]=len(val_set)
xlsx.save(r'.\\data\\result.xlsx')
xlsx.close()  # 确保文件被关闭

print("=== 优化前评估 ===")
print("优化前在训练集的平均f1值：",avg_scores_before[2])

# 优化器将只调整 TargetModule 中声明为 dspy.Module 的属性
optimized_pipeline = optimizer.compile(
    pipeline_module_targeted, 
    # num_trials=20,
    trainset=trainset,
    valset=val_set)
# dspy.inspect_history(n=3)

# 优化后，获取优化后的 predictor 实例
summarizer_opt = optimized_pipeline.summarizer

#提取优化后提示词
summary_prompt_opt=summarizer_opt.signature.instructions
# 打开文件并写入内容（自动关闭文件），更新summary提示词
with open( r'.\data\prompt\detail_summary_prompt_opt'+str(epoch_num_cv)+'.txt', "w", encoding="utf-8") as file12:
    file12.write(summary_prompt_opt)
    
scores_after = [
    evalution_metric(example, optimized_pipeline(**example.inputs(), similar_threshold=DEFAULT_SIMILAR_THRESHOLD))
    for example in val_set
    ]

# 过滤掉 -1
scores_after_filtered = [s for s in scores_after if s != -1]
# avg_after = sum(scores_after_filtered) / len(scores_after_filtered)
# 转为 numpy 数组计算平均值
scores_after_np = np.array(scores_after_filtered)
avg_scores_after = np.mean(scores_after_np, axis=0)



print("\n=== 优化后评估（ ===")
print("优化后在训练集的平均f1值:", avg_scores_after[2])

end_time=time.time()
cost_time=(end_time-start_time)/3600
print(f'***优化时长为：{cost_time:.2f}小时,当前并行计算线程数为：{num_cores}***')

xlsx=openpyxl.load_workbook(r'.\\data\\result.xlsx')
sheet1=xlsx['CV']
line=(epoch_num_cv)*3
sheet1['C'+str(line)]=avg_scores_after[0]
sheet1['D'+str(line)]=avg_scores_after[1]
sheet1['E'+str(line)]=avg_scores_after[2]
sheet1['F'+str(line)]=cost_time
sheet1['G'+str(line)]=DEFAULT_SIMILAR_THRESHOLD
sheet1['H'+str(line)]=len(trainset)
sheet1['I'+str(line)]=len(val_set)
xlsx.save(r'.\\data\\result.xlsx')
xlsx.close()  # 确保文件被关闭

# 保存优化后的模型（按示例程序风格）
try:
    # optimized_pipeline.save(".\\data\\dspy_opted_model\\optimized_pipeline_partial_dual_lm_avg5.json")
    optimized_pipeline.save(".\\data\\dspy_opted_model\\opted_model\\optimized_dspy_True"+str(epoch_num_cv), save_program=True)
    print("\n已保存优化后的模型：optimized_dspy_True")
except Exception as e1:
    print("\n保存优化模型save_program=True时遇到问题：", e1)
    try:
        optimized_pipeline.save(".\\data\\dspy_opted_model\\opted_model\\optimized_dspy_False"+str(epoch_num_cv)+'.pkl', save_program=False)
        print("\n保存save_program=False成功")
    except Exception as e2:   
        print("保存优化后模型save_program=False均不成功！！！",e2)

# 返回或打印最终关键指标
print("\n=== 最终总结 ===")
print("优化前平均f1:", avg_scores_before[2])
print("优化后平均f1:", avg_scores_after[2])
print('当前运行的单元格是In[2]，已经完成。')

# In[3] optuna优化超参数
# ==============================
# Optuna 超参数优化部分（无全局变量）
# ==============================
import optuna
start_time_optuna=time.time()
def objective_optuna(trial):
    # 建议搜索范围根据任务调整，例如 [0.6, 0.9]
    similar_threshold = trial.suggest_float('similar_threshold_value', 0.6, 0.9)

    total_f1 = 0.0
    valid_count = 0
    for example in val_set:  # 使用验证集评估！
        # === MODIFIED FOR OPTUNA ===
        pred = optimized_pipeline(**example.inputs(), similar_threshold=similar_threshold)
        score = my_metric(example, pred)
        if score != -1:
            total_f1 += score
            valid_count += 1

    avg_f1 = total_f1 / valid_count if valid_count > 0 else 0.0
    return -avg_f1  # Optuna minimize，所以取负

print("\n开始使用 Optuna 优化 similar_threshold_value（基于 DSPy 优化后的 pipeline）...")

study = optuna.create_study(direction='minimize')
study.optimize(objective_optuna, n_trials=30)

best_threshold = study.best_params['similar_threshold_value']
best_f1 = -study.best_value

end_time_optuna=time.time()
cost_time_optuna=(end_time_optuna-start_time_optuna)/3600
print("\n Optuna 优化完成！")
print(f"最佳 similar_threshold_value: {best_threshold:.4f}")
print(f"对应验证集平均 F1: {best_f1:.4f}")

    
try:
    if final_threshold is not None:  # 检查a存在且不为None
       final_threshold[epoch_num_cv]=best_threshold
except NameError:
    final_threshold={}
    final_threshold[epoch_num_cv]=best_threshold
print('当前运行的单元格是In[3]，已经完成。')    

final_threshold={1:0.748927032487498,
                 2:0.682200337168919,
                 3:0.765677072185444,
                 4:0.766117280557596,
                 5:0.761023331551931}

# In[4]  计算验证集得分

# =============================================================================
# 采用加载保存的优化模型计算得分
# =============================================================================
# scores_opt = [opt_metric(example, opt_pipeline_module_targeted(**example.inputs())) for example in val_set]

print(f'验证集包含的数据共{len(val_set)}条。')
# =============================================================================
# 直接采用优化模型计算得分
# =============================================================================

scores_opt = [
    evalution_metric(example, optimized_pipeline(**example.inputs(), similar_threshold=final_threshold[epoch_num_cv]))
    for example in val_set
    ]

# 过滤掉 -1
scores_opt_filtered = [s for s in scores_opt if s != -1]

# 转为 numpy 数组计算平均值
scores_opt_np = np.array(scores_opt_filtered)
avg_scores_opt = np.mean(scores_opt_np, axis=0)

print(f'***第{epoch_num_cv}轮交叉验证，验证集指标平均值：precision是{avg_scores_opt[0]},recall是{avg_scores_opt[1]},f1是{avg_scores_opt[2]}。')

xlsx=openpyxl.load_workbook(r'.\\data\\result.xlsx')
sheet1=xlsx['CV']
line=epoch_num_cv*3+1
sheet1['C'+str(line)]=avg_scores_opt[0]
sheet1['D'+str(line)]=avg_scores_opt[1]
sheet1['E'+str(line)]=avg_scores_opt[2]
sheet1['F'+str(line)]=cost_time_optuna
sheet1['G'+str(line)]=final_threshold[epoch_num_cv]
sheet1['H'+str(line)]=len(trainset)
sheet1['I'+str(line)]=len(val_set)
xlsx.save(r'.\\data\\result.xlsx')
xlsx.close()  # 确保文件被关闭
print('当前运行的单元格是In[4]，已经完成。')

# In[5.1]测试集的生成，本单元格运行一次即可
# =============================================================================
# 单元格[5.1]~[5.4]是计算测试集的评价指标,In[5.1]是测试集的生成，本单元格运行一次即可
# =============================================================================
dict_candidate_summary_pickle=r'.\data\dspy_opted_model\dict_candidate_summary.pkl'
with open(dict_candidate_summary_pickle,'rb') as f_obj1:
     dict_candidate_summary=pickle.load(f_obj1) 
     
dict_candidate_chain_pickle=r'.\data\dspy_opted_model\dict_candidate_chain.pkl'
with open(dict_candidate_chain_pickle,'rb') as f_obj2:
     dict_candidate_chain=pickle.load(f_obj2) 
     
embedding_dict_candidate_pickle=r'.\data\dspy_opted_model\embedding_dict_candidate.pkl'
with open(embedding_dict_candidate_pickle,'rb') as f_obj3:
     embedding_dict_candidate=pickle.load(f_obj3)              

try:
    list_query_test_pickle=r'.\data\dspy_opted_model\list_query_test.pkl'
    with open(list_query_test_pickle,'rb') as f_obj12:
         list_query_test=pickle.load(f_obj12)    
    assert len(list_query_test) == num_testset,f"加载的测试集案例数量为 {len(list_query_test)} 与预期 {num_testset} 不符"
except (FileNotFoundError, EOFError, pickle.UnpicklingError, AssertionError) as e_list_query_test:    
    print('加载测试集时出错',e_list_query_test)

    if isinstance(e_list_query_test, FileNotFoundError):
        print('测试集文件未找到，需要重新生成。')
        list_query_test=[]
    elif isinstance(e_list_query_test, AssertionError):
        print(f'测试集加载的查询案例不够，需要补充生成{num_testset-len(list_query_test)}条案例数据。')
        
    add_num_query_test=num_testset-len(list_query_test)
    for _ in range(add_num_query_test): 
         dict1={}
         dict_chain={}
         list_candidate=random.sample(list(embedding_dict_candidate.keys()), num_candidate_case)  
         for k_dict in list_candidate:

# =============================================================================
#         #从每个候选案例的行为链词嵌入字典中随机抽取键构建dict1
# =============================================================================
             dict_summaries = embedding_dict_candidate[k_dict]
             list_summary_label=random.sample(list(dict_summaries.keys()), min(num_summary,len(dict_summaries.keys())))  
        
             dict1[k_dict]=copy.deepcopy(list_summary_label)
             sub_dict_chain={}
             for key_list in list_summary_label:
          
             # chain_extracted=action_chain_extractor(summary=dict_summaries[key_list])   #获取某条summary的行为链
             # chain_gen = getattr(chain_extracted, "action_chain", "")
             # sub_dict_chain[key_list]=chain_gen
                 try:
#对如下语句进行修改，用summary代替chain生成query ****************************************************            
                     # sub_dict_chain[key_list]=dict_candidate_chain[k_dict][key_list]
                     sub_dict_chain[key_list]=dict_candidate_summary[k_dict][key_list]
                 except Exception as e:
                     print(f'在第{k_dict}个候选案例的第{key_list}摘要调用对应行为链出错',e)
                     print(f'第{k_dict}个候选案例的摘要键列表为{list_summary_label}，行为链键列表为{list(dict_candidate_chain[k_dict])}')
             dict_chain[k_dict]=sub_dict_chain            
         dict_query=generate_query_and_validate(composer, judge, dict_chain, dict1, max_retries=5)
         if dict_query['text']=='failed':
              continue
         list_query_test.append(dict_query) 

    list_query_test_pickle=r'.\data\dspy_opted_model\list_query_test.pkl'  
    with open(list_query_test_pickle,'wb') as f_obj13:
         pickle.dump(list_query_test, f_obj13)           
                  
testset = []                             #dataset是全部数据的集合
for query_test in list_query_test:   
    example = dspy.Example(dict_query=query_test, label=query_test['label']).with_inputs("dict_query")
    testset.append(example)
print('当前运行的单元格是In[5.1]，已经完成。')    

# In[5.2]调用测试集进行计算，本单元格运行一次即可
#加载最初优化前的提示词
file_path_summary = r'.\data\prompt\detail_summary_prompt_opt0-2.txt'    #读取优化后的提示词文本文件
with open(file_path_summary, 'r', encoding='utf-8') as file1:       
      detail_summary_prompt_1 = file1.read()      
PROMPT_SUMMARY_1 = detail_summary_prompt_1 
summarizer_1 = dspy.Predict(make_signature("text", "summary", PROMPT_SUMMARY_1)) 
   
summarizer_2 = copy.deepcopy(summarizer_1)

print('当前运行的单元格是In[5.2]，已经完成。')  

# In[5.3]可以修改本单元格式的epoch_num_cv值，使用最优的模型和final_threshold计算指标值，
# =============================================================================
# 本单元格可以修改epoch_num_cv值，多次运行，用测试集计算交叉验证不同折保存模型的指标
# =============================================================================

#num_selected可以修改！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！
print(f'当前测试集中数据共有{len(testset)}个。')
num_selected=5

# loaded_pipeline_2 是第epoch_num_cv折交叉验证优化后的模型
loaded_pipeline_2 = TargetedPipelineModule(summarizer=summarizer_2)
loaded_pipeline_2.load(".\\data\\dspy_opted_model\\opted_model\\optimized_dspy_False"+str(num_selected)+'.pkl')

scores_test_2 = [
    evalution_metric(example, loaded_pipeline_2(**example.inputs(), similar_threshold=final_threshold[num_selected]))
    for example in testset
    ]

# 过滤掉 -1
scores_test_2_filtered = [s for s in scores_test_2 if s != -1]

# 转为 numpy 数组计算平均值
scores_test_2_np = np.array(scores_test_2_filtered)
avg_scores_test_2 = np.mean(scores_test_2_np, axis=0)

print(f'***测试集在第{num_selected}折交叉验证优化后模型上的指标平均值：precision是{avg_scores_test_2[0]},recall是{avg_scores_test_2[1]},f1是{avg_scores_test_2[2]}。')

xlsx=openpyxl.load_workbook(r'.\\data\\result.xlsx')
sheet1=xlsx['test']
line=num_selected+1
sheet1['C'+str(line)]=avg_scores_test_2[0]
sheet1['D'+str(line)]=avg_scores_test_2[1]
sheet1['E'+str(line)]=avg_scores_test_2[2]
sheet1['F'+str(line)]=final_threshold[num_selected]   
sheet1['G'+str(line)]=len(testset) 
sheet1['H'+str(line)]=num_selected     
xlsx.save(r'.\\data\\result.xlsx')
xlsx.close()  # 确保文件被关闭
print('当前运行的单元格是In[5.3]，已经完成。')  

# In[5.4] 采用原始模型计算测试集得分，本单元格运行一次即可

#loaded_pipeline_1是优化前的原始模型
loaded_pipeline_1 = TargetedPipelineModule(summarizer=summarizer_1)

scores_test_1 = [
    evalution_metric(example, loaded_pipeline_1(**example.inputs(), similar_threshold=DEFAULT_SIMILAR_THRESHOLD))
    for example in testset
    ]

# 过滤掉 -1
scores_test_1_filtered = [s for s in scores_test_1 if s != -1]

# 转为 numpy 数组计算平均值
scores_test_1_np = np.array(scores_test_1_filtered)
avg_scores_test_1 = np.mean(scores_test_1_np, axis=0)

print(f'***测试集在原始模型上的指标平均值：precision是{avg_scores_test_1[0]},recall是{avg_scores_test_1[1]},f1是{avg_scores_test_1[2]}。')

xlsx=openpyxl.load_workbook(r'.\\data\\result.xlsx')
sheet1=xlsx['test']
line=7
sheet1['C'+str(line)]=avg_scores_test_1[0]
sheet1['D'+str(line)]=avg_scores_test_1[1]
sheet1['E'+str(line)]=avg_scores_test_1[2]
sheet1['F'+str(line)]=DEFAULT_SIMILAR_THRESHOLD
sheet1['G'+str(line)]=len(testset) 
sheet1['H'+str(line)]=num_selected   
xlsx.save(r'.\\data\\result.xlsx')
xlsx.close()  # 确保文件被关闭

print('当前运行的单元格是In[5.4]，已经完成。')  

# In[5.5] 提示词未优化，参数已经优化，计算测试集得分，本单元格运行一次即可

#loaded_pipeline_1是优化前的原始模型
loaded_pipeline_1_1 = copy.deepcopy(loaded_pipeline_1)

scores_test_1_1 = [
    evalution_metric(example, loaded_pipeline_1_1(**example.inputs(), similar_threshold=final_threshold[num_selected]))
    for example in testset
    ]

# 过滤掉 -1
scores_test_1_1_filtered = [s for s in scores_test_1_1 if s != -1]

# 转为 numpy 数组计算平均值
scores_test_1_1_np = np.array(scores_test_1_1_filtered)
avg_scores_test_1_1 = np.mean(scores_test_1_1_np, axis=0)

print(f'***测试集在提示词未优化，参数已经优化的模型上的指标平均值：precision是{avg_scores_test_1_1[0]},recall是{avg_scores_test_1_1[1]},f1是{avg_scores_test_1_1[2]}。')

xlsx=openpyxl.load_workbook(r'.\\data\\result.xlsx')
sheet1=xlsx['test']
line=8
sheet1['C'+str(line)]=avg_scores_test_1_1[0]
sheet1['D'+str(line)]=avg_scores_test_1_1[1]
sheet1['E'+str(line)]=avg_scores_test_1_1[2]
sheet1['F'+str(line)]=final_threshold[num_selected]
sheet1['G'+str(line)]=len(testset) 
sheet1['H'+str(line)]=num_selected   
xlsx.save(r'.\\data\\result.xlsx')
xlsx.close()  # 确保文件被关闭

print('当前运行的单元格是In[5.5]，已经完成。')  

# In[5.6] 提示词已经优化，参数未优化，计算测试集得分，本单元格运行一次即可

# loaded_pipeline_2 是第epoch_num_cv折交叉验证优化后的模型
loaded_pipeline_2_2=copy.deepcopy(loaded_pipeline_2)

scores_test_2_2 = [
    evalution_metric(example, loaded_pipeline_2_2(**example.inputs(), similar_threshold=DEFAULT_SIMILAR_THRESHOLD))
    for example in testset
    ]

# 过滤掉 -1
scores_test_2_2_filtered = [s for s in scores_test_2_2 if s != -1]

# 转为 numpy 数组计算平均值
scores_test_2_2_np = np.array(scores_test_2_2_filtered)
avg_scores_test_2_2 = np.mean(scores_test_2_2_np, axis=0)

print(f'***测试集在提示词已经优化，参数未优化的模型上的指标平均值：precision是{avg_scores_test_2_2[0]},recall是{avg_scores_test_2_2[1]},f1是{avg_scores_test_2_2[2]}。')

xlsx=openpyxl.load_workbook(r'.\\data\\result.xlsx')
sheet1=xlsx['test']
line=9
sheet1['C'+str(line)]=avg_scores_test_2_2[0]
sheet1['D'+str(line)]=avg_scores_test_2_2[1]
sheet1['E'+str(line)]=avg_scores_test_2_2[2]
sheet1['F'+str(line)]=DEFAULT_SIMILAR_THRESHOLD 
sheet1['G'+str(line)]=len(testset) 
sheet1['H'+str(line)]=num_selected        
xlsx.save(r'.\\data\\result.xlsx')
xlsx.close()  # 确保文件被关闭

print('当前运行的单元格是In[5.6]，已经完成。')  

# In[6]生成裁判规则

# import docx
# from langchain.prompts import (
#     ChatPromptTemplate,  
#     SystemMessagePromptTemplate,
#     HumanMessagePromptTemplate)
# from langchain_openai import ChatOpenAI
# from langchain.output_parsers import CommaSeparatedListOutputParser

# file_path = r'.\data\prompt\query.txt'    #读取查询案例文本
# with open(file_path, 'r', encoding='utf-8') as file:       
#      query_text = file.read()

# output_parser=CommaSeparatedListOutputParser()  #实例化输出解析器
     
# xlsx_candidation = openpyxl.load_workbook(r'.\data\候选案例.xlsx')
# sheet_candidate = xlsx_candidation['Sheet1']

# llm_rule = ChatOpenAI(model='qwen-plus',base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
#                    temperature=0.0,top_p=0.1,
#                    openai_api_key=sheet_api_key['B'+str(2)].value)
# system_template = "你是一个法律专家"
# system_message_template = SystemMessagePromptTemplate.from_template(system_template)
# file_path_judgement = r'.\data\prompt\judgement_rule_prompt.txt'    #读取提示词文本文件
# with open(file_path_judgement, 'r', encoding='utf-8') as file_judgement:       
#      judgement_prompt = file_judgement.read()
# human_message_prompt_judgement = HumanMessagePromptTemplate.from_template(judgement_prompt)
# chat_prompt_judgement = ChatPromptTemplate.from_messages([system_message_template, human_message_prompt_judgement])
# chain_judgement = chat_prompt_judgement|llm_rule|output_parser   # 新版langchain导入的链  

# query_dict_4_induce={"text":query_text,"label":{1:[1,2]}}
# query_example = dspy.Example(dict_query=query_dict_4_induce, label=query_dict_4_induce['label']).with_inputs("dict_query")
# dict_pre=loaded_pipeline_2(**query_example.inputs(),similar_threshold=final_threshold[num_selected])
# query_chains=dict_pre.query_chains
# query2candidate=dict_pre.query2candidate
     
# dict_summary_candidate_pickle=r'.\data\dspy_opted_model\dict_candidate_summary.pkl'
# with open(dict_summary_candidate_pickle, 'rb') as f_obj8:
#     dict_summary_candidate = pickle.load(f_obj8)     

# doc=docx.Document()     
# for q3 in query2candidate.keys():
#     query_input=query_chains['chain'][q3]
#     k3=query2candidate[q3][0]
#     reason_input=sheet_candidate['L'+str(k3)].value
#     candidate_input=dict_summary_candidate[k3][query2candidate[q3][1]]
#     response=chain_judgement.invoke({'query':query_input,'candidate':candidate_input,'reason':reason_input})[2]
#     name_candidate=sheet_candidate['A'+str(k3)].value
#     doc.add_paragraph(f'{q3}、查询案例中：{query_input}根据第{k3}个判例，得出的裁判规则是：\n{response}\n参考的案例是：{ name_candidate}\n')
# doc.save(r'.\data\output\result.docx')

# print('当前运行的单元格是In[6]，已经完成。')  