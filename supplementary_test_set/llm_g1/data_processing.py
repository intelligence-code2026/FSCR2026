# 本程序用于从判决书中提取案件的事实部分
# This program is designed to extract the factual portion of cases from court judgments.


import openpyxl
from langchain_openai import ChatOpenAI
from langchain.output_parsers import CommaSeparatedListOutputParser
from langchain.prompts import (
    ChatPromptTemplate,
    SystemMessagePromptTemplate,
    HumanMessagePromptTemplate)

xlsx_api_key=openpyxl.load_workbook(r'..\..\api_key.xlsx')
sheet_api_key=xlsx_api_key['Sheet1']



system_template = "你是一个法律专家"
system_message_template = SystemMessagePromptTemplate.from_template(system_template)
output_parser=CommaSeparatedListOutputParser()  #实例化输出解析器

# 初始化 LLM 和链
llm = ChatOpenAI(model='qwen-plus',base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
                   temperature=0.0,top_p=0.1,
                   openai_api_key=sheet_api_key['B'+str(2)].value)

llm_ds = ChatOpenAI(model='deepseek-chat',base_url="https://api.deepseek.com",
                   temperature=0.0,top_p=0.1,
                   openai_api_key=sheet_api_key['B'+str(7)].value)

file_path_data = r'.\data\prompt\text_data.txt'    #读取提示词文本文件
with open(file_path_data, 'r', encoding='utf-8') as file:       
     data_prompt = file.read()
human_message_prompt_data = HumanMessagePromptTemplate.from_template(data_prompt)
chat_prompt_summary = ChatPromptTemplate.from_messages([system_message_template, human_message_prompt_data])
chain_data = chat_prompt_summary|llm_ds|output_parser   # 新版langchain导入的链


xlsx_data = openpyxl.load_workbook(r'.\data\candidate_case2.xlsx')
sheet_data = xlsx_data['Sheet1']
sheet2=xlsx_data['Sheet2']
for i in list(range(sheet_data.max_row))[:]:
    text_data = sheet_data['C' + str(i+2)].value
    detail_text = ''.join(chain_data.invoke({'input':text_data}))
    sheet2['B'+str(i+2)]=detail_text

xlsx_data.save(r'.\data\candidate_case2.xlsx')
xlsx_data.close()  # 确保文件被关闭
