# 计算真实查询案例与每个候选案例的整体余弦相似度
# Calculate the overall cosine similarity between the real query case and each candidate case
from openai import OpenAI
import pickle
import openpyxl
import sys
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
list_similar=[]
# In[0]
num_query=1

file_path = '.\\data\\query'+str(num_query)+'.txt'    #读取查询案例文本
with open(file_path, 'r', encoding='utf-8') as file:       
      query_text = file.read()
      
# 读取 API Key
# Load API Key
try:
    xlsx_api_key = openpyxl.load_workbook(r'..\api_key.xlsx')
    sheet_api_key = xlsx_api_key['Sheet1']
    api_key_deepseek = sheet_api_key['B7'].value
    api_key = sheet_api_key['B' + str(2)].value
except Exception as e:
    print('未找到llm的api_key！！！')
    print(f"错误详情: {e}")
    sys.exit(1)


client = OpenAI(
   api_key=sheet_api_key['B'+str(2)].value,  
   base_url="https://dashscope.aliyuncs.com/compatible-mode/v1"  )      
      
def embedding_vector(sentence:str,client=client)->list:     
    completion = client.embeddings.create(
        model="text-embedding-v3",
        input=sentence,
        dimensions=1024,
        encoding_format="float"
    )
    return  completion.data[0].embedding  

def similiar_score(embedding_vec_query, embedding_vec_candidate)->float:
    # 确保输入是numpy数组（如果你的输入已经是数组，可以省略这步）
    vec1 = np.array(embedding_vec_query).reshape(1, -1)  # 强制转为2D: (1, n)
    vec2 = np.array(embedding_vec_candidate).reshape(1, -1)  # 强制转为2D: (1, n)
    
    # 计算余弦相似度（结果是一个1x1矩阵）
    similarity_matrix = cosine_similarity(vec1, vec2)
    
    # 返回标量值（提取矩阵中的单个数值）
    return similarity_matrix[0][0]

dict_case_text_pickle=r'.\data\dspy_opted_model\dict_case_text.pkl'
with open(dict_case_text_pickle,'rb') as f_obj9:
     dict_case_text=pickle.load(f_obj9) 
     
embedding_query= embedding_vector(query_text) 


xlsx=openpyxl.load_workbook(r'.\\data\\result.xlsx')
sheet1=xlsx['similar']

dict_result={1:'B',2:'D',3:'F'}
sub_list_similar=[]

for key in dict_case_text.keys():
    embedding_candidate=embedding_vector(dict_case_text[key]) 
    similiar_folat=similiar_score(embedding_query, embedding_candidate)
    sub_list_similar.append(similiar_folat)
    sheet1[dict_result[num_query]+str(key)]=similiar_folat       
xlsx.save(r'.\\data\\result.xlsx')
xlsx.close()  # 确保文件被关闭    
list_similar.append(sub_list_similar)    
# In[1]
array_similar=np.array(list_similar) 
global_min = min(np.min(d) for d in array_similar)
global_max = max(np.max(d) for d in array_similar)

bin_edges = np.linspace(global_min, global_max, 11)
histograms = []

cumulative_histograms = []
cumulative_frequencies = []

for i, data in enumerate(array_similar):   
    counts, _ = np.histogram(data, bins=bin_edges)
    histograms.append(counts)
    
    cum_counts=np.cumsum(counts)
    cumulative_histograms.append(cum_counts)
    
    cum_freq=cum_counts/array_similar.shape[1]
    cumulative_frequencies.append(cum_freq)
    
bin_midpoints = (bin_edges[:-1] + bin_edges[1:]) / 2    

xlsx=openpyxl.load_workbook(r'.\\data\\result.xlsx')
sheet1=xlsx['histograms']
dict_histograms={1:'B',2:'C',3:'D'}
for i in range(len(bin_midpoints)):
    sheet1['A'+str(i+2)]=bin_midpoints[i]
for j in range(len(cumulative_frequencies)):
    for k in range(len(cumulative_frequencies[j])):
        sheet1[dict_histograms[j+1]+str(k+2)]=cumulative_frequencies[j][k] 
xlsx.save(r'.\\data\\result.xlsx')
xlsx.close()         